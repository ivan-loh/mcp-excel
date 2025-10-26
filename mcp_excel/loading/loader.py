import re
from pathlib import Path
from typing import Optional
import pandas as pd
import duckdb

from ..models import SheetOverride, TableMeta, MergeHandlingConfig
from ..utils.naming import TableRegistry
from .formats.manager import FormatManager
from .analyzer import ExcelAnalyzer
from ..utils import log


class ExcelLoader:
    def __init__(self, conn: duckdb.DuckDBPyConnection, registry: TableRegistry):
        self.conn = conn
        self.registry = registry
        self.format_manager = FormatManager()
        self._ensure_excel_extension()

    def _ensure_excel_extension(self):
        try:
            self.conn.execute("INSTALL excel")
            self.conn.execute("LOAD excel")
        except Exception:
            pass

    def load_sheet(
        self,
        file: Path,
        relpath: str,
        sheet: str,
        alias: str,
        override: Optional[SheetOverride] = None,
    ) -> list[TableMeta]:
        effective_override = override
        structure_info = None

        if override and override.auto_detect and file.suffix.lower() in ['.xlsx', '.xlsm']:
            try:
                analyzer = ExcelAnalyzer()
                structure_info = analyzer.analyze_structure(file, sheet)

                effective_override = self._merge_override_with_detection(override, structure_info)

                log.info("structure_analysis_applied",
                        file=str(file),
                        sheet=sheet,
                        merged_cells=len(structure_info.merged_ranges),
                        hidden_rows=len(structure_info.hidden_rows),
                        header_row=structure_info.header_row,
                        num_tables=structure_info.num_tables)
            except Exception as e:
                log.warn("structure_analysis_failed", file=str(file), sheet=sheet, error=str(e))
                effective_override = override

        if effective_override:
            self._validate_override_options(effective_override, structure_info)

        if effective_override and structure_info and structure_info.num_tables > 1:
            return self._load_multi_table(file, relpath, sheet, alias, effective_override, structure_info)

        table_name = self.registry.register(alias, relpath, sheet)

        if effective_override:
            single_meta = self._load_assisted(file, relpath, sheet, table_name, alias, effective_override)
        else:
            single_meta = self._load_raw(file, relpath, sheet, table_name, alias)

        return [single_meta]

    def _load_multi_table(
        self,
        file: Path,
        relpath: str,
        sheet: str,
        alias: str,
        override: SheetOverride,
        structure_info
    ) -> list[TableMeta]:
        table_metas = []

        if override.table_range:
            log.info("multi_table_override_with_range",
                    file=str(file), sheet=sheet, range=override.table_range)
            base_table_name = self.registry.register(alias, relpath, sheet)
            single_meta = self._load_assisted(file, relpath, sheet, base_table_name, alias, override)
            return [single_meta]

        if override.extract_table is not None:
            if override.extract_table < 0 or override.extract_table >= len(structure_info.table_ranges):
                log.warn("invalid_extract_table_index",
                        extract_table=override.extract_table,
                        num_tables=len(structure_info.table_ranges))
                table_idx = 0
            else:
                table_idx = override.extract_table

            table_info = structure_info.table_ranges[table_idx]
            table_name_suffix = f"_table{table_idx}" if len(structure_info.table_ranges) > 1 else ""
            table_name = self.registry.register(alias, relpath, f"{sheet}{table_name_suffix}")

            meta = self._load_table_range(file, relpath, sheet, table_name, alias, override, table_info)
            return [meta]

        for idx, table_info in enumerate(structure_info.table_ranges):
            table_name_suffix = f"_table{idx}" if len(structure_info.table_ranges) > 1 else ""
            table_name = self.registry.register(alias, relpath, f"{sheet}{table_name_suffix}")

            log.info("loading_table_from_range",
                    file=str(file), sheet=sheet, table_index=idx,
                    start_row=table_info['start_row'], end_row=table_info['end_row'])

            try:
                meta = self._load_table_range(file, relpath, sheet, table_name, alias, override, table_info)
                table_metas.append(meta)
            except Exception as e:
                log.warn("table_load_failed", table_index=idx, error=str(e))
                continue

        return table_metas if table_metas else [self._load_raw(file, relpath, sheet,
                                                               self.registry.register(alias, relpath, sheet), alias)]

    def _load_table_range(
        self,
        file: Path,
        relpath: str,
        sheet: str,
        table_name: str,
        alias: str,
        override: SheetOverride,
        table_info: dict
    ) -> TableMeta:
        start_row = table_info['start_row']
        end_row = table_info['end_row']
        start_col = table_info['start_col']
        end_col = table_info['end_col']

        from openpyxl.utils import get_column_letter
        start_col_letter = get_column_letter(start_col)
        end_col_letter = get_column_letter(end_col)

        range_spec = f"{start_col_letter}{start_row}:{end_col_letter}{end_row}"

        range_override = SheetOverride(
            skip_rows=0,
            header_rows=1 if table_info.get('has_header', True) else 0,
            skip_footer=0,
            range=range_spec,
            drop_regex=override.drop_regex,
            column_renames=override.column_renames,
            type_hints=override.type_hints,
            unpivot=override.unpivot,
            auto_detect=False,
            merge_handling=override.merge_handling,
            include_hidden=override.include_hidden,
            locale=override.locale,
            drop_conditions=override.drop_conditions,
            extract_table=None,
            table_range=None
        )

        return self._load_assisted(file, relpath, sheet, table_name, alias, range_override)

    def _merge_override_with_detection(self, override: SheetOverride, structure_info) -> SheetOverride:
        merged = SheetOverride(
            skip_rows=override.skip_rows if override.skip_rows > 0 else structure_info.suggested_skip_rows,
            header_rows=override.header_rows,
            skip_footer=override.skip_footer if override.skip_footer > 0 else structure_info.suggested_skip_footer,
            range=override.range,
            drop_regex=override.drop_regex,
            column_renames=override.column_renames,
            type_hints=override.type_hints,
            unpivot=override.unpivot,
            auto_detect=override.auto_detect,
            merge_handling=override.merge_handling if override.merge_handling else (
                MergeHandlingConfig() if len(structure_info.merged_ranges) > 0 else None
            ),
            include_hidden=override.include_hidden,
            locale=override.locale,
            drop_conditions=override.drop_conditions,
            extract_table=override.extract_table,
            table_range=override.table_range
        )

        return merged

    def _load_raw(self, file: Path, relpath: str, sheet: str, table_name: str, alias: str) -> TableMeta:
        try:
            if file.suffix.lower() not in ['.xlsx', '.xlsm']:
                df = self.format_manager.load_file(file, sheet, {'normalize': False})

                import hashlib
                temp_table = f"temp_{hashlib.md5(table_name.encode()).hexdigest()[:8]}"
                self.conn.register(temp_table, df)

                self.conn.execute(f"""
                    CREATE OR REPLACE VIEW "{table_name}" AS
                    SELECT * FROM {temp_table}
                """)
                est_rows = len(df)
            else:
                self.conn.execute(f"""
                    CREATE OR REPLACE VIEW "{table_name}" AS
                    SELECT * FROM read_xlsx(
                        '{file}',
                        sheet='{sheet}',
                        header=false,
                        all_varchar=true
                    )
                """)
                count_result = self.conn.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()
                est_rows = count_result[0] if count_result else 0

            return TableMeta(
                table_name=table_name,
                file=str(file),
                relpath=relpath,
                sheet=sheet,
                mode="RAW",
                mtime=file.stat().st_mtime,
                alias=alias,
                est_rows=est_rows,
            )
        except Exception as e:
            error_msg = str(e)
            suggestion = self._get_error_suggestion(error_msg, "RAW")
            raise RuntimeError(f"Failed to load {file}:{sheet} in RAW mode: {error_msg}{suggestion}")

    def _load_assisted(
        self, file: Path, relpath: str, sheet: str, table_name: str, alias: str, override: SheetOverride
    ) -> TableMeta:
        try:
            use_format_manager = (
                file.suffix.lower() not in ['.xlsx', '.xlsm'] or
                (not override.include_hidden) or
                override.merge_handling is not None
            )

            if use_format_manager:
                options = {
                    'skip_rows': override.skip_rows,
                    'header_rows': override.header_rows,
                    'skip_footer': override.skip_footer,
                    'normalize': True,
                    'merge_strategy': override.merge_handling.strategy if override.merge_handling else 'fill',
                    'ignore_hidden': override.include_hidden == False,
                    'range': override.range if override.range else None,
                }
                if override.locale:
                    self.format_manager.normalizer.set_locale({
                        'locale': override.locale.locale,
                        'decimal_separator': override.locale.decimal_separator,
                        'thousands_separator': override.locale.thousands_separator,
                        'currency_symbols': override.locale.currency_symbols,
                        'auto_detect': override.locale.auto_detect
                    })
                df = self.format_manager.load_file(file, sheet, options)

                if override.drop_regex and len(df.columns) > 0:
                    first_col = df.columns[0]
                    df = df[~df[first_col].astype(str).str.match(override.drop_regex, na=False)]

                if override.drop_conditions:
                    df = self._apply_drop_conditions(df, override.drop_conditions)

                if override.column_renames:
                    df = df.rename(columns=override.column_renames)

                if override.type_hints:
                    df = self._apply_type_hints(df, override.type_hints)

                if override.unpivot:
                    df = self._apply_unpivot(df, override.unpivot)
            elif override.header_rows > 1:
                df = self._load_multirow_header(file, sheet, override)
            else:
                has_header = override.header_rows > 0
                range_clause = f", range='{override.range}'" if override.range else ""

                df = self.conn.execute(f"""
                    SELECT * FROM read_xlsx(
                        '{file}',
                        sheet='{sheet}',
                        header={has_header},
                        all_varchar=true
                        {range_clause}
                    )
                """).df()

                if override.skip_rows > 0 and not override.range and override.header_rows <= 1:
                    df = df.iloc[override.skip_rows:]

                if override.skip_footer > 0:
                    df = df.iloc[:-override.skip_footer]

                if override.drop_regex:
                    if len(df.columns) > 0:
                        first_col = df.columns[0]
                        pattern = override.drop_regex
                        df = df[~df[first_col].astype(str).str.match(pattern, na=False)]

                if override.drop_conditions:
                    df = self._apply_drop_conditions(df, override.drop_conditions)

                if override.column_renames:
                    df = df.rename(columns=override.column_renames)

                if override.type_hints:
                    df = self._apply_type_hints(df, override.type_hints)

                if override.unpivot:
                    df = self._apply_unpivot(df, override.unpivot)

            import hashlib
            temp_view = f"temp_{hashlib.md5(table_name.encode()).hexdigest()[:8]}"
            self.conn.register(temp_view, df)
            self.conn.execute(f"""
                CREATE OR REPLACE VIEW "{table_name}" AS
                SELECT * FROM {temp_view}
            """)

            return TableMeta(
                table_name=table_name,
                file=str(file),
                relpath=relpath,
                sheet=sheet,
                mode="ASSISTED",
                mtime=file.stat().st_mtime,
                alias=alias,
                est_rows=len(df),
            )
        except Exception as e:
            error_msg = str(e)
            suggestion = self._get_error_suggestion(error_msg, "ASSISTED")
            raise RuntimeError(f"Failed to load {file}:{sheet} in ASSISTED mode: {error_msg}{suggestion}")

    def _load_multirow_header(self, file: Path, sheet: str, override: SheetOverride) -> pd.DataFrame:
        df_raw = self.conn.execute(f"""
            SELECT * FROM read_xlsx(
                '{file}',
                sheet='{sheet}',
                header=false,
                all_varchar=true
                {f", range='{override.range}'" if override.range else ""}
            )
        """).df()

        if override.skip_rows > 0:
            df_raw = df_raw.iloc[override.skip_rows:]

        header_rows = df_raw.iloc[:override.header_rows]
        data_rows = df_raw.iloc[override.header_rows:]

        new_columns = []
        for col_idx in range(len(header_rows.columns)):
            col_parts = []
            for row_idx in range(len(header_rows)):
                val = str(header_rows.iloc[row_idx, col_idx])
                if val and val != "nan":
                    col_parts.append(val)

            if col_parts:
                new_col_name = "__".join(col_parts)
            else:
                new_col_name = f"col_{col_idx}"

            new_columns.append(new_col_name)

        data_rows.columns = new_columns
        data_rows = data_rows.reset_index(drop=True)

        return data_rows

    def _apply_type_hints(self, df: pd.DataFrame, type_hints: dict[str, str]) -> pd.DataFrame:
        for col_name, type_hint in type_hints.items():
            if col_name not in df.columns:
                continue

            type_upper = type_hint.upper()
            integer_types = ("INT", "BIGINT", "SMALLINT")
            numeric_types = ("DECIMAL", "NUMERIC", "DOUBLE", "FLOAT")

            if any(t in type_upper for t in integer_types):
                df[col_name] = pd.to_numeric(df[col_name], errors="coerce").astype("Int64")
            elif any(t in type_upper for t in numeric_types):
                df[col_name] = pd.to_numeric(df[col_name], errors="coerce")
            elif "DATE" in type_upper:
                df[col_name] = pd.to_datetime(df[col_name], errors="coerce")
            elif "BOOL" in type_upper:
                df[col_name] = df[col_name].astype(str).str.lower().isin(["true", "1", "yes", "y"])

        return df

    def _apply_unpivot(self, df: pd.DataFrame, unpivot_config: dict) -> pd.DataFrame:
        id_vars = unpivot_config.get("id_vars", [])
        value_vars = unpivot_config.get("value_vars", [])
        var_name = unpivot_config.get("var_name", "variable")
        value_name = unpivot_config.get("value_name", "value")

        if not value_vars:
            value_vars = [col for col in df.columns if col not in id_vars]

        return df.melt(id_vars=id_vars, value_vars=value_vars, var_name=var_name, value_name=value_name)

    def _apply_drop_conditions(self, df: pd.DataFrame, drop_conditions: list[dict]) -> pd.DataFrame:
        if not drop_conditions or len(df) == 0:
            return df

        mask = pd.Series([True] * len(df), index=df.index)
        total_dropped = 0

        for condition in drop_conditions:
            col_name = condition.get('column')
            if not col_name:
                log.warn("drop_condition_missing_column", condition=str(condition))
                continue

            if col_name not in df.columns:
                log.warn("drop_condition_column_not_found", column=col_name, available_columns=list(df.columns))
                continue

            condition_mask = pd.Series([False] * len(df), index=df.index)

            if 'regex' in condition:
                pattern = condition['regex']
                try:
                    condition_mask = df[col_name].astype(str).str.match(pattern, na=False)
                    dropped = int(condition_mask.sum())
                    if dropped > 0:
                        log.info("drop_condition_regex_applied", column=col_name, pattern=pattern, rows_dropped=dropped)
                except Exception as e:
                    log.warn("drop_condition_regex_failed", column=col_name, pattern=pattern, error=str(e))

            elif 'equals' in condition:
                value = condition['equals']
                condition_mask = df[col_name] == value
                dropped = int(condition_mask.sum())
                if dropped > 0:
                    log.info("drop_condition_equals_applied", column=col_name, value=value, rows_dropped=dropped)

            elif 'is_null' in condition:
                if condition['is_null']:
                    condition_mask = df[col_name].isna()
                    dropped = int(condition_mask.sum())
                    if dropped > 0:
                        log.info("drop_condition_null_applied", column=col_name, rows_dropped=dropped)
                else:
                    condition_mask = df[col_name].notna()
                    dropped = int(condition_mask.sum())
                    if dropped > 0:
                        log.info("drop_condition_not_null_applied", column=col_name, rows_dropped=dropped)

            else:
                log.warn("drop_condition_unknown_operator", condition=str(condition))

            mask = mask & ~condition_mask
            total_dropped += int(condition_mask.sum())

        result_df = df[mask].reset_index(drop=True)

        if total_dropped > 0:
            log.info("drop_conditions_completed", original_rows=len(df), dropped_rows=total_dropped, remaining_rows=len(result_df))

        return result_df

    def _validate_override_options(self, override: SheetOverride, structure_info: Optional[Any] = None):
        if override.extract_table is not None and override.table_range:
            log.warn("conflicting_options",
                    message="Both extract_table and table_range specified, table_range will take precedence")

        if override.auto_detect and override.skip_rows > 0:
            log.info("auto_detect_with_skip_rows",
                    message="auto_detect enabled with skip_rows, manual skip_rows will override detection")

        if structure_info and hasattr(structure_info, 'header_confidence'):
            if structure_info.header_confidence < 0.3:
                log.warn("low_header_confidence",
                        confidence=structure_info.header_confidence,
                        suggestion="Consider manual 'header_rows' configuration for better results")

        if override.drop_conditions and override.drop_regex:
            log.info("both_drop_methods_used",
                    message="Both drop_conditions and drop_regex specified, both will be applied")

    def _get_error_suggestion(self, error_msg: str, mode: str) -> str:
        error_lower = error_msg.lower()
        suggestions = []

        if "column" in error_lower and "mismatch" in error_lower:
            suggestions.append("Try adding 'skip_rows' to skip header rows")
            suggestions.append("Or use 'drop_regex' to filter problematic rows")

        if "header" in error_lower or "column name" in error_lower:
            suggestions.append("Consider using 'header_rows: 0' or 'header_rows: 2' to adjust header detection")
            suggestions.append("Use 'column_renames' to fix column names")

        if "row" in error_lower and ("empty" in error_lower or "null" in error_lower):
            suggestions.append("Use 'skip_footer' to remove trailing empty rows")
            suggestions.append("Or 'drop_regex' to filter specific rows")

        if "type" in error_lower or "convert" in error_lower or "cast" in error_lower:
            suggestions.append("Use 'type_hints' to specify column types explicitly")
            suggestions.append("Consider loading in RAW mode first to inspect the data")

        if "range" in error_lower:
            suggestions.append("Check that 'range' parameter uses valid Excel notation (e.g., 'A1:F100')")

        if mode == "RAW" and not suggestions:
            suggestions.append("Try ASSISTED mode with overrides to handle messy data")
            suggestions.append("Use 'skip_rows' and 'skip_footer' to exclude problematic rows")

        if suggestions:
            return "\n\nSuggestions:\n- " + "\n- ".join(suggestions)
        return ""

    def get_sheet_names(self, file: Path) -> list[str]:
        if file.suffix.lower() not in ['.xlsx', '.xlsm', '.xls']:
            return self.format_manager.get_sheets(file)

        try:
            result = self.conn.execute(f"""
                SELECT sheet_name FROM st_read('{file}')
            """).fetchall()
            return [row[0] for row in result]
        except Exception:
            try:
                return self.format_manager.get_sheets(file)
            except:
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
                    sheets = wb.sheetnames
                    wb.close()
                    return sheets
                except Exception as e:
                    raise RuntimeError(f"Failed to read sheet names from {file}: {e}")

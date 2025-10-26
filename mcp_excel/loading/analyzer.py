from pathlib import Path
from typing import Optional
from collections import OrderedDict
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from ..models import StructureInfo
from ..utils import log


class LRUCache:
    def __init__(self, maxsize: int = 128):
        self.cache: OrderedDict[str, StructureInfo] = OrderedDict()
        self.maxsize = maxsize

    def get(self, key: str) -> Optional[StructureInfo]:
        if key not in self.cache:
            return None
        self.cache.move_to_end(key)
        return self.cache[key]

    def put(self, key: str, value: StructureInfo):
        if key in self.cache:
            self.cache.move_to_end(key)
        self.cache[key] = value
        if len(self.cache) > self.maxsize:
            self.cache.popitem(last=False)

    def __contains__(self, key: str) -> bool:
        return key in self.cache

    def __len__(self) -> int:
        return len(self.cache)

    def clear(self):
        self.cache.clear()


class ExcelAnalyzer:
    def __init__(self, cache_size: int = 128):
        self._cache = LRUCache(maxsize=cache_size)

    def analyze_structure(self, file_path: Path, sheet: str) -> StructureInfo:
        mtime = file_path.stat().st_mtime
        cache_key = f"{file_path}:{sheet}:{mtime}"

        cached_result = self._cache.get(cache_key)
        if cached_result:
            log.info("structure_cache_hit", file=str(file_path), sheet=sheet, cache_size=len(self._cache))
            return cached_result

        log.info("structure_analysis_started", file=str(file_path), sheet=sheet)

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
            ws = wb[sheet] if sheet else wb.active

            merged_ranges = self._detect_merged_cells(ws)
            hidden_rows = self._detect_hidden_rows(ws)
            hidden_columns = self._detect_hidden_columns(ws)
            data_region = self._detect_data_region(ws, hidden_rows, hidden_columns)
            header_info = self._detect_headers(ws, data_region, merged_ranges)
            metadata_info = self._detect_metadata(ws, data_region, header_info)
            locale_info = self._detect_locale(ws, data_region)
            tables = self._detect_multiple_tables(ws, data_region, header_info)
            blank_rows = self._detect_blank_rows(ws, data_region)

            structure_info = StructureInfo(
                data_start_row=data_region['start_row'],
                data_end_row=data_region['end_row'],
                data_start_col=data_region['start_col'],
                data_end_col=data_region['end_col'],
                header_row=header_info['header_row'],
                header_rows_count=header_info['header_rows_count'],
                header_confidence=header_info['confidence'],
                metadata_rows=metadata_info['rows'],
                metadata_type=metadata_info['type'],
                merged_ranges=merged_ranges,
                merged_in_headers=self._has_merged_in_range(merged_ranges, header_info['header_row'], header_info['header_row'] + header_info['header_rows_count'] - 1) if header_info['header_row'] else False,
                merged_in_data=len(merged_ranges) > 0,
                hidden_rows=hidden_rows,
                hidden_columns=hidden_columns,
                detected_locale=locale_info['locale'],
                decimal_separator=locale_info['decimal_separator'],
                thousands_separator=locale_info['thousands_separator'],
                num_tables=len(tables),
                table_ranges=tables,
                blank_rows=blank_rows,
                inconsistent_columns=data_region.get('inconsistent_columns', False),
                has_formulas=False,
                suggested_skip_rows=metadata_info['suggested_skip'],
                suggested_skip_footer=0,
                suggested_overrides={}
            )

            wb.close()

            self._cache.put(cache_key, structure_info)

            log.info("structure_detected",
                    cache_size=len(self._cache),
                    data_region=f"{structure_info.data_start_row}:{structure_info.data_end_row}",
                    header_row=structure_info.header_row,
                    header_confidence=structure_info.header_confidence,
                    merged_cells_count=len(merged_ranges),
                    hidden_rows_count=len(hidden_rows),
                    detected_locale=structure_info.detected_locale,
                    num_tables=structure_info.num_tables)

            return structure_info

        except Exception as e:
            log.error("structure_analysis_failed", file=str(file_path), sheet=sheet, error=str(e))
            return self._create_default_structure_info()

    def _detect_merged_cells(self, ws: Worksheet) -> list[tuple[int, int, int, int]]:
        merged = []
        for merged_range in ws.merged_cells.ranges:
            merged.append((
                merged_range.min_row,
                merged_range.min_col,
                merged_range.max_row,
                merged_range.max_col
            ))
        return merged

    def _detect_hidden_rows(self, ws: Worksheet) -> list[int]:
        hidden = []
        for row_num in range(1, ws.max_row + 1):
            if ws.row_dimensions[row_num].hidden:
                hidden.append(row_num)
        return hidden

    def _detect_hidden_columns(self, ws: Worksheet) -> list[int]:
        hidden = []
        for col_num in range(1, ws.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            if ws.column_dimensions[col_letter].hidden:
                hidden.append(col_num)
        return hidden

    def _detect_data_region(self, ws: Worksheet, hidden_rows: list[int], hidden_columns: list[int]) -> dict:
        start_row = None
        start_col = None
        end_row = None
        end_col = None

        for row_idx in range(1, ws.max_row + 1):
            if row_idx in hidden_rows:
                continue
            for col_idx in range(1, ws.max_column + 1):
                if col_idx in hidden_columns:
                    continue
                cell = ws.cell(row_idx, col_idx)
                if cell.value is not None:
                    if start_row is None:
                        start_row = row_idx
                        start_col = col_idx
                    end_row = row_idx
                    if end_col is None or col_idx > end_col:
                        end_col = col_idx
                    if start_col is None or col_idx < start_col:
                        start_col = col_idx

        if start_row is None:
            start_row = 1
            start_col = 1
            end_row = 1
            end_col = 1

        return {
            'start_row': start_row,
            'end_row': end_row,
            'start_col': start_col,
            'end_col': end_col,
            'inconsistent_columns': False
        }

    def _detect_headers(self, ws: Worksheet, data_region: dict, merged_ranges: list) -> dict:
        start_row = data_region['start_row']

        for row_idx in range(start_row, min(start_row + 10, data_region['end_row'] + 1)):
            row_cells = [
                ws.cell(row_idx, col).value
                for col in range(data_region['start_col'], data_region['end_col'] + 1)
            ]

            non_empty = [c for c in row_cells if c is not None]
            if not non_empty:
                continue

            all_strings = all(isinstance(c, str) for c in non_empty)

            if row_idx < data_region['end_row']:
                next_row_cells = [
                    ws.cell(row_idx + 1, col).value
                    for col in range(data_region['start_col'], data_region['end_col'] + 1)
                ]
                has_numbers_below = any(
                    isinstance(c, (int, float)) for c in next_row_cells if c is not None
                )
            else:
                has_numbers_below = False

            if all_strings and (has_numbers_below or row_idx == data_region['end_row']):
                return {
                    'header_row': row_idx,
                    'header_rows_count': 1,
                    'confidence': 0.9
                }

        return {
            'header_row': None,
            'header_rows_count': 0,
            'confidence': 0.0
        }

    def _detect_metadata(self, ws: Worksheet, data_region: dict, header_info: dict) -> dict:
        metadata_rows = []
        metadata_type = 'unknown'

        if header_info['header_row'] and header_info['header_row'] > data_region['start_row']:
            for row_idx in range(data_region['start_row'], header_info['header_row']):
                metadata_rows.append(row_idx)
            metadata_type = 'title'

        suggested_skip = len(metadata_rows)

        return {
            'rows': metadata_rows,
            'type': metadata_type,
            'suggested_skip': suggested_skip
        }

    def _detect_locale(self, ws: Worksheet, data_region: dict) -> dict:
        number_formats = []
        sample_values = []

        for row in range(data_region['start_row'], min(data_region['start_row'] + 50, data_region['end_row'] + 1)):
            for col in range(data_region['start_col'], min(data_region['start_col'] + 20, data_region['end_col'] + 1)):
                cell = ws.cell(row, col)
                if isinstance(cell.value, (int, float)):
                    number_formats.append(cell.number_format)
                elif isinstance(cell.value, str):
                    sample_values.append(cell.value)

        european_format_indicators = ['#.##0,00', '#,##0.00_-', '0.00_-', '[$€-*] #,##0.00']
        has_european_format = any(
            any(pattern in str(fmt) for pattern in european_format_indicators)
            for fmt in number_formats
        )

        import re
        has_comma_decimal = sum(1 for s in sample_values if re.search(r'\d,\d{2}(?!\d)', s))
        has_dot_decimal = sum(1 for s in sample_values if re.search(r'\d\.\d{2}(?!\d)', s))
        has_dot_thousands = sum(1 for s in sample_values if re.search(r'\d\.\d{3}', s))
        has_comma_thousands = sum(1 for s in sample_values if re.search(r'\d,\d{3}', s))

        if has_european_format or (has_comma_decimal > has_dot_decimal and has_dot_thousands > has_comma_thousands):
            return {
                'locale': 'de_DE',
                'decimal_separator': ',',
                'thousands_separator': '.'
            }
        else:
            return {
                'locale': 'en_US',
                'decimal_separator': '.',
                'thousands_separator': ','
            }

    def _detect_multiple_tables(self, ws: Worksheet, data_region: dict, header_info: dict) -> list[dict]:
        tables = []
        blank_rows = self._detect_blank_rows(ws, data_region)

        blank_row_groups = self._group_consecutive_blank_rows(blank_rows)

        significant_separators = [group for group in blank_row_groups if len(group) >= 2]

        if not significant_separators:
            if header_info['header_row']:
                tables.append({
                    'start_row': header_info['header_row'],
                    'end_row': data_region['end_row'],
                    'start_col': data_region['start_col'],
                    'end_col': data_region['end_col'],
                    'has_header': True,
                    'header_row': header_info['header_row'],
                    'confidence': header_info['confidence']
                })
            return tables

        table_sections = self._split_by_separators(
            data_region['start_row'],
            data_region['end_row'],
            significant_separators
        )

        for section_start, section_end in table_sections:
            if section_start > section_end:
                continue

            section_header = self._detect_section_header(
                ws, section_start, section_end,
                data_region['start_col'], data_region['end_col']
            )

            if section_header['found']:
                title_rows = self._detect_title_rows(
                    ws, section_start, section_header['header_row'],
                    data_region['start_col'], data_region['end_col']
                )

                table_start_col, table_end_col = self._detect_table_width(
                    ws, section_header['header_row'], section_end,
                    data_region['start_col'], data_region['end_col']
                )

                tables.append({
                    'start_row': section_header['header_row'],
                    'end_row': section_end,
                    'start_col': table_start_col,
                    'end_col': table_end_col,
                    'has_header': True,
                    'header_row': section_header['header_row'],
                    'confidence': section_header['confidence'],
                    'title_rows': title_rows
                })

        if not tables and header_info['header_row']:
            tables.append({
                'start_row': header_info['header_row'],
                'end_row': data_region['end_row'],
                'start_col': data_region['start_col'],
                'end_col': data_region['end_col'],
                'has_header': True,
                'header_row': header_info['header_row'],
                'confidence': header_info['confidence']
            })

        return tables

    def _detect_blank_rows(self, ws: Worksheet, data_region: dict) -> list[int]:
        blank = []
        for row_idx in range(data_region['start_row'], data_region['end_row'] + 1):
            row_cells = [
                ws.cell(row_idx, col).value
                for col in range(data_region['start_col'], data_region['end_col'] + 1)
            ]
            if all(c is None for c in row_cells):
                blank.append(row_idx)
        return blank

    def _group_consecutive_blank_rows(self, blank_rows: list[int]) -> list[list[int]]:
        if not blank_rows:
            return []

        groups = []
        current_group = [blank_rows[0]]

        for i in range(1, len(blank_rows)):
            if blank_rows[i] == blank_rows[i-1] + 1:
                current_group.append(blank_rows[i])
            else:
                groups.append(current_group)
                current_group = [blank_rows[i]]

        groups.append(current_group)
        return groups

    def _split_by_separators(self, start_row: int, end_row: int, separators: list[list[int]]) -> list[tuple[int, int]]:
        sections = []
        current_start = start_row

        for separator_group in sorted(separators, key=lambda g: g[0]):
            separator_start = separator_group[0]
            separator_end = separator_group[-1]

            if current_start < separator_start:
                sections.append((current_start, separator_start - 1))

            current_start = separator_end + 1

        if current_start <= end_row:
            sections.append((current_start, end_row))

        return sections

    def _detect_section_header(self, ws: Worksheet, section_start: int, section_end: int,
                               start_col: int, end_col: int) -> dict:
        for row_idx in range(section_start, min(section_start + 10, section_end + 1)):
            row_cells = [
                ws.cell(row_idx, col).value
                for col in range(start_col, end_col + 1)
            ]

            non_empty = [c for c in row_cells if c is not None]
            if not non_empty:
                continue

            all_strings = all(isinstance(c, str) for c in non_empty)

            if row_idx < section_end:
                next_row_cells = [
                    ws.cell(row_idx + 1, col).value
                    for col in range(start_col, end_col + 1)
                ]
                has_numbers_below = any(
                    isinstance(c, (int, float)) for c in next_row_cells if c is not None
                )
            else:
                has_numbers_below = False

            if all_strings and (has_numbers_below or row_idx == section_end):
                confidence = 0.9 if has_numbers_below else 0.5
                return {
                    'found': True,
                    'header_row': row_idx,
                    'confidence': confidence
                }

        return {
            'found': False,
            'header_row': None,
            'confidence': 0.0
        }

    def _detect_title_rows(self, ws: Worksheet, section_start: int, header_row: int,
                          start_col: int, end_col: int) -> list[int]:
        title_rows = []

        for row_idx in range(section_start, header_row):
            row_cells = [
                ws.cell(row_idx, col).value
                for col in range(start_col, end_col + 1)
            ]

            non_empty = [c for c in row_cells if c is not None]
            if len(non_empty) > 0:
                title_rows.append(row_idx)

        return title_rows

    def _detect_table_width(self, ws: Worksheet, header_row: int, end_row: int,
                           start_col: int, end_col: int) -> tuple[int, int]:
        min_col = end_col
        max_col = start_col

        for row_idx in range(header_row, min(header_row + 100, end_row + 1)):
            for col_idx in range(start_col, end_col + 1):
                cell = ws.cell(row_idx, col_idx)
                if cell.value is not None:
                    min_col = min(min_col, col_idx)
                    max_col = max(max_col, col_idx)

        return (min_col, max_col) if min_col <= max_col else (start_col, end_col)

    def _has_merged_in_range(self, merged_ranges: list, start_row: int, end_row: int) -> bool:
        for min_row, min_col, max_row, max_col in merged_ranges:
            if start_row <= min_row <= end_row or start_row <= max_row <= end_row:
                return True
        return False

    def _create_default_structure_info(self) -> StructureInfo:
        return StructureInfo(
            data_start_row=1,
            data_end_row=1,
            data_start_col=1,
            data_end_col=1,
            header_row=None,
            header_rows_count=0,
            header_confidence=0.0,
            metadata_rows=[],
            metadata_type='unknown',
            merged_ranges=[],
            merged_in_headers=False,
            merged_in_data=False,
            hidden_rows=[],
            hidden_columns=[],
            detected_locale='en_US',
            decimal_separator='.',
            thousands_separator=',',
            num_tables=0,
            table_ranges=[],
            blank_rows=[],
            inconsistent_columns=False,
            has_formulas=False,
            suggested_skip_rows=0,
            suggested_skip_footer=0,
            suggested_overrides={}
        )

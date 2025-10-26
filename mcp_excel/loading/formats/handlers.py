from abc import ABC, abstractmethod
from pathlib import Path
import pandas as pd
import numpy as np
from typing import Optional, List
from dataclasses import dataclass, field
import csv

from ...exceptions import FormatDetectionError, FileError
from ...utils import log

@dataclass
class ParseOptions:
    encoding: str = 'utf-8'
    skip_rows: int = 0
    header_rows: int = 1
    skip_footer: int = 0
    na_values: List[str] = field(default_factory=list)
    parse_dates: bool = True
    chunk_size: Optional[int] = None
    data_only: bool = True
    preserve_formatting: bool = False
    handle_merged_cells: str = 'fill'
    ignore_hidden: bool = False
    max_rows: Optional[int] = None
    merge_strategy: str = 'fill'
    range: Optional[str] = None

class FormatHandler(ABC):
    @abstractmethod
    def can_handle(self, format_type: str) -> bool:
        pass

    @abstractmethod
    def parse(self, file_path: Path, sheet: Optional[str], options: ParseOptions) -> pd.DataFrame:
        pass

    @abstractmethod
    def get_sheets(self, file_path: Path) -> List[str]:
        pass

    @abstractmethod
    def validate(self, file_path: Path) -> tuple[bool, Optional[str]]:
        pass

class XLSXHandler(FormatHandler):
    def can_handle(self, format_type: str) -> bool:
        return format_type in ['xlsx', 'xlsm']

    def parse(self, file_path: Path, sheet: Optional[str], options: ParseOptions) -> pd.DataFrame:
        try:
            import openpyxl
        except ImportError:
            return pd.read_excel(
                file_path,
                sheet_name=sheet,
                skiprows=options.skip_rows,
                skipfooter=options.skip_footer,
                na_values=options.na_values,
                engine='openpyxl'
            )

        read_only = options.merge_strategy == 'skip'
        wb = openpyxl.load_workbook(
            file_path,
            read_only=read_only,
            data_only=options.data_only,
            keep_links=False
        )

        ws = wb[sheet] if sheet else wb.active

        if options.merge_strategy != 'skip' and not read_only:
            ws = self._handle_merged_cells(ws, options.merge_strategy)

        data = []

        if options.range:
            rows_iter = ws[options.range]
            if not isinstance(rows_iter[0], tuple):
                rows_iter = [rows_iter]

            for row in rows_iter:
                if not isinstance(row, tuple):
                    row = [row]
                row_values = [cell.value for cell in row]
                data.append(row_values)
        else:
            for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
                if options.ignore_hidden and hasattr(ws.row_dimensions[row_idx], 'hidden') and ws.row_dimensions[row_idx].hidden:
                    continue

                row_values = []
                for cell in row:
                    if options.ignore_hidden and hasattr(ws.column_dimensions[cell.column_letter], 'hidden') and ws.column_dimensions[cell.column_letter].hidden:
                        continue
                    row_values.append(cell.value)

                data.append(row_values)

        wb.close()

        if not data:
            return pd.DataFrame()

        if options.skip_rows > 0:
            data = data[options.skip_rows:]

        if options.skip_footer > 0:
            data = data[:-options.skip_footer]

        if options.header_rows > 0 and len(data) > options.header_rows:
            headers = data[:options.header_rows]
            data = data[options.header_rows:]

            if options.header_rows == 1:
                columns = [str(h) if h else f'col_{i}' for i, h in enumerate(headers[0])]
            else:
                columns = []
                for col_idx in range(len(headers[0])):
                    col_parts = []
                    for row in headers:
                        if col_idx < len(row) and row[col_idx]:
                            col_parts.append(str(row[col_idx]))
                    columns.append('_'.join(col_parts) if col_parts else f'col_{col_idx}')
        else:
            columns = [f'col_{i}' for i in range(len(data[0]) if data else 0)]

        df = pd.DataFrame(data, columns=columns)
        df = self._clean_excel_data(df, options)

        return df

    def _handle_merged_cells(self, ws, strategy: str):
        import openpyxl

        merged_ranges = list(ws.merged_cells.ranges)

        for merged_range in merged_ranges:
            top_left_value = ws.cell(
                merged_range.min_row,
                merged_range.min_col
            ).value

            ws.unmerge_cells(str(merged_range))

            if strategy == 'fill':
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        ws.cell(row, col).value = top_left_value

        return ws

    def _clean_excel_data(self, df: pd.DataFrame, options: ParseOptions) -> pd.DataFrame:
        excel_errors = ['#DIV/0!', '#N/A', '#NAME?', '#NULL!', '#NUM!', '#REF!', '#VALUE!']
        df = df.replace(excel_errors, np.nan)

        if options.na_values:
            df = df.replace(options.na_values, np.nan)

        return df

    def get_sheets(self, file_path: Path) -> List[str]:
        errors_encountered = []
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return sheets
        except PermissionError as e:
            raise FileError(
                f"Permission denied reading {file_path}",
                file_path=str(file_path),
                operation="get_sheets",
                data={"error": str(e), "handler": "XLSXHandler"}
            )
        except FileNotFoundError as e:
            raise FileError(
                f"File not found: {file_path}",
                file_path=str(file_path),
                operation="get_sheets",
                data={"error": str(e)}
            )
        except Exception as e:
            errors_encountered.append(f"openpyxl: {type(e).__name__}: {str(e)}")
            log.debug("openpyxl_failed_trying_pandas", file=str(file_path), error=str(e))

            try:
                xl_file = pd.ExcelFile(file_path, engine='openpyxl')
                sheets = xl_file.sheet_names
                xl_file.close()
                log.info("pandas_fallback_success", file=str(file_path))
                return sheets
            except (PermissionError, FileNotFoundError):
                raise
            except Exception as fallback_error:
                errors_encountered.append(f"pandas: {type(fallback_error).__name__}: {str(fallback_error)}")
                raise FormatDetectionError(
                    f"Failed to read sheets from {file_path.name}",
                    file_path=str(file_path),
                    attempted_formats=["openpyxl", "pandas"],
                    data={"errors": errors_encountered}
                )

    def validate(self, file_path: Path) -> tuple[bool, Optional[str]]:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            wb.close()
            return True, None
        except Exception as e:
            return False, str(e)

class XLSHandler(FormatHandler):
    def can_handle(self, format_type: str) -> bool:
        return format_type == 'xls'

    def parse(self, file_path: Path, sheet: Optional[str], options: ParseOptions) -> pd.DataFrame:
        try:
            df = pd.read_excel(
                file_path,
                sheet_name=sheet if sheet else 0,
                skiprows=options.skip_rows,
                skipfooter=options.skip_footer,
                na_values=options.na_values,
                engine='xlrd'
            )
            return df
        except ImportError:
            return pd.read_excel(
                file_path,
                sheet_name=sheet if sheet else 0,
                skiprows=options.skip_rows,
                skipfooter=options.skip_footer,
                na_values=options.na_values
            )

    def get_sheets(self, file_path: Path) -> List[str]:
        try:
            xl_file = pd.ExcelFile(file_path)
            sheets = xl_file.sheet_names
            xl_file.close()
            return sheets
        except (PermissionError, FileNotFoundError):
            raise
        except Exception as e:
            log.warn("xls_get_sheets_failed", file=str(file_path), error=str(e))
            return ['Sheet1']

    def validate(self, file_path: Path) -> tuple[bool, Optional[str]]:
        try:
            xl_file = pd.ExcelFile(file_path)
            xl_file.close()
            return True, None
        except Exception as e:
            return False, str(e)

class CSVHandler(FormatHandler):
    def can_handle(self, format_type: str) -> bool:
        return format_type in ['csv', 'tsv']

    def parse(self, file_path: Path, sheet: Optional[str], options: ParseOptions) -> pd.DataFrame:
        encoding = self._detect_encoding(file_path) if options.encoding == 'utf-8' else options.encoding
        delimiter = self._detect_delimiter(file_path, encoding)

        try:
            df = pd.read_csv(
                file_path,
                encoding=encoding,
                delimiter=delimiter,
                skiprows=options.skip_rows,
                skipfooter=options.skip_footer,
                nrows=options.max_rows,
                na_values=options.na_values or ['', 'NA', 'N/A', 'null', 'NULL', '#N/A'],
                parse_dates=options.parse_dates,
                chunksize=options.chunk_size,
                engine='python' if options.skip_footer > 0 else 'c'
            )
            return df
        except UnicodeDecodeError:
            encodings_to_try = ['latin-1', 'windows-1252', 'iso-8859-1']
            errors = []
            for enc in encodings_to_try:
                try:
                    df = pd.read_csv(
                        file_path,
                        encoding=enc,
                        delimiter=delimiter,
                        skiprows=options.skip_rows,
                        skipfooter=options.skip_footer,
                        nrows=options.max_rows,
                        na_values=options.na_values or ['', 'NA', 'N/A', 'null', 'NULL', '#N/A'],
                        parse_dates=options.parse_dates,
                        engine='python' if options.skip_footer > 0 else 'c'
                    )
                    log.info("csv_encoding_fallback_success", file=str(file_path), encoding=enc)
                    return df
                except UnicodeDecodeError as e:
                    errors.append(f"{enc}: UnicodeDecodeError")
                    log.debug("csv_encoding_failed", encoding=enc, error=str(e))
                    continue
                except Exception as e:
                    errors.append(f"{enc}: {type(e).__name__}: {e}")
                    log.debug("csv_read_failed", encoding=enc, error=str(e))
                    continue
            raise FormatDetectionError(
                f"Failed to read CSV with any encoding: {file_path}",
                file_path=str(file_path),
                attempted_formats=[f"csv[{enc}]" for enc in [encoding] + encodings_to_try],
                data={"errors": errors}
            )

    def _detect_delimiter(self, file_path: Path, encoding: str) -> str:
        try:
            with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                sample = f.read(8192)

            sniffer = csv.Sniffer()
            dialect = sniffer.sniff(sample)
            log.debug("delimiter_sniffed", delimiter=repr(dialect.delimiter))
            return dialect.delimiter
        except csv.Error as e:
            log.debug("delimiter_sniff_failed", error=str(e), reason="csv_error")
        except Exception as e:
            log.debug("delimiter_sniff_failed", error=str(e), error_type=type(e).__name__)
            with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                lines = f.readlines()[:10]

            delimiters = [',', '\t', ';', '|']
            delimiter_scores = {}

            for delim in delimiters:
                counts = [line.count(delim) for line in lines if line.strip()]
                if counts and counts[0] > 0:
                    avg_count = sum(counts) / len(counts)
                    variance = sum((c - avg_count) ** 2 for c in counts) / len(counts) if len(counts) > 1 else 0
                    delimiter_scores[delim] = avg_count / (variance + 1)

            if delimiter_scores:
                return max(delimiter_scores, key=delimiter_scores.get)

            return ','

    def _detect_encoding(self, file_path: Path) -> str:
        with open(file_path, 'rb') as f:
            raw_data = f.read(10000)

        if raw_data.startswith(b'\xef\xbb\xbf'):
            return 'utf-8-sig'

        if raw_data.startswith(b'\xff\xfe') or raw_data.startswith(b'\xfe\xff'):
            return 'utf-16'

        try:
            raw_data.decode('utf-8')
            return 'utf-8'
        except UnicodeDecodeError:
            pass

        try:
            raw_data.decode('latin-1')
            return 'latin-1'
        except UnicodeDecodeError:
            pass

        return 'windows-1252'

    def get_sheets(self, file_path: Path) -> List[str]:
        return ['Sheet1']

    def validate(self, file_path: Path) -> tuple[bool, Optional[str]]:
        try:
            encoding = self._detect_encoding(file_path)
            with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                f.read(1024)
            return True, None
        except Exception as e:
            return False, str(e)
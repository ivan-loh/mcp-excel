import pytest
from pathlib import Path
import openpyxl

from mcp_excel.models import SheetOverride, StructureInfo
from mcp_excel.loading.loader import ExcelLoader
from mcp_excel.utils.naming import TableRegistry
import duckdb

pytestmark = pytest.mark.unit


def test_drop_conditions_missing_column_field(temp_dir, loader):
    file_path = temp_dir / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Name"])
    ws.append(["A"])
    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        drop_conditions=[
            {"regex": "test"}
        ]
    )
    metas = loader.load_sheet(file_path, "test.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').df()

    assert len(result) == 1


def test_drop_conditions_unknown_operator(temp_dir, loader):
    file_path = temp_dir / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Name"])
    ws.append(["A"])
    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        drop_conditions=[
            {"column": "Name", "unknown_op": "value"}
        ]
    )
    metas = loader.load_sheet(file_path, "test.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').df()

    assert len(result) == 1


def test_validate_override_options_low_confidence():
    conn = duckdb.connect()
    registry = TableRegistry()
    loader = ExcelLoader(conn, registry)

    structure_info = StructureInfo(
        data_start_row=1,
        data_end_row=10,
        data_start_col=1,
        data_end_col=5,
        header_row=1,
        header_rows_count=1,
        header_confidence=0.2,
        metadata_rows=[],
        metadata_type="none",
        merged_ranges=[],
        merged_in_headers=False,
        merged_in_data=False,
        hidden_rows=[],
        hidden_columns=[],
        detected_locale="en_US",
        decimal_separator=".",
        thousands_separator=",",
        num_tables=1,
        table_ranges=[],
        blank_rows=[],
        inconsistent_columns=False,
        has_formulas=False,
        suggested_skip_rows=0,
        suggested_skip_footer=0,
        suggested_overrides={}
    )

    override = SheetOverride(auto_detect=True)

    loader._validate_override_options(override, structure_info)

    conn.close()


def test_drop_conditions_with_type_hints(temp_dir, loader):
    file_path = temp_dir / "types.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Name", "Amount"])
    ws.append(["A", "100.5"])
    ws.append(["B", "200.5"])
    ws.append(["TOTAL", "301.0"])
    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        type_hints={"Amount": "DECIMAL(10,2)"},
        drop_conditions=[
            {"column": "Name", "regex": "^TOTAL"}
        ]
    )
    metas = loader.load_sheet(file_path, "types.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').df()

    assert len(result) == 2


def test_complex_drop_conditions_combination(temp_dir, loader):
    file_path = temp_dir / "complex.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Product", "Amount", "Status", "Category"])
    ws.append(["Widget A", "100", "ACTIVE", "Cat1"])
    ws.append(["TOTAL", "100", "", ""])
    ws.append(["Widget B", None, "ACTIVE", "Cat2"])
    ws.append(["Widget C", "300", "DELETED", "Cat1"])
    ws.append(["SUBTOTAL", "300", "", ""])
    ws.append(["Widget D", "400", "ACTIVE", None])
    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        drop_conditions=[
            {"column": "Product", "regex": "^(TOTAL|SUBTOTAL)"},
            {"column": "Status", "equals": "DELETED"},
            {"column": "Amount", "is_null": True},
            {"column": "Category", "is_null": True}
        ]
    )
    metas = loader.load_sheet(file_path, "complex.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').df()

    assert len(result) == 1
    assert result["Product"].values[0] == "Widget A"


def test_conflicting_options_validation(temp_dir, loader):
    file_path = temp_dir / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["A"])
    ws.append(["1"])
    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        extract_table=0,
        table_range="A1:B2"
    )

    metas = loader.load_sheet(file_path, "test.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    assert meta is not None

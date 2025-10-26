import pytest
from pathlib import Path
import pandas as pd
import openpyxl

from mcp_excel.types import SheetOverride

pytestmark = pytest.mark.unit


def test_drop_conditions_regex_basic(temp_dir, loader):
    file_path = temp_dir / "totals.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Product", "Amount", "Status"])
    ws.append(["Widget A", "100", "ACTIVE"])
    ws.append(["Widget B", "200", "ACTIVE"])
    ws.append(["TOTAL", "300", ""])
    ws.append(["Widget C", "150", "DELETED"])
    ws.append(["Widget D", "250", "ACTIVE"])
    ws.append(["SUBTOTAL", "400", ""])
    ws.append(["Grand Total", "700", ""])
    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        drop_conditions=[
            {"column": "Product", "regex": "^(TOTAL|SUBTOTAL|Grand Total)"}
        ]
    )
    metas = loader.load_sheet(file_path, "totals.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').df()

    assert len(result) == 4
    products = result["Product"].tolist()
    assert "Widget A" in products
    assert "Widget B" in products
    assert "Widget C" in products
    assert "Widget D" in products
    assert "TOTAL" not in products
    assert "SUBTOTAL" not in products
    assert "Grand Total" not in products


def test_drop_conditions_equals(temp_dir, loader):
    file_path = temp_dir / "status.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Product", "Status"])
    ws.append(["Widget A", "ACTIVE"])
    ws.append(["Widget B", "DELETED"])
    ws.append(["Widget C", "ACTIVE"])
    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        drop_conditions=[
            {"column": "Status", "equals": "DELETED"}
        ]
    )
    metas = loader.load_sheet(file_path, "status.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').df()

    assert len(result) == 2
    assert "DELETED" not in result["Status"].values


def test_drop_conditions_is_null(temp_dir, loader):
    file_path = temp_dir / "nulls.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Name", "Value"])
    ws.append(["Item A", "100"])
    ws.append(["Item B", None])
    ws.append(["Item C", "300"])
    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        drop_conditions=[
            {"column": "Value", "is_null": True}
        ]
    )
    metas = loader.load_sheet(file_path, "nulls.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').df()

    assert len(result) == 2
    names = result["Name"].tolist()
    assert "Item A" in names
    assert "Item C" in names
    assert "Item B" not in names


def test_drop_conditions_multiple(temp_dir, loader):
    file_path = temp_dir / "multiple.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Product", "Amount", "Status"])
    ws.append(["Widget A", "100", "ACTIVE"])
    ws.append(["Widget B", "200", "ACTIVE"])
    ws.append(["TOTAL", "300", ""])
    ws.append(["Widget C", "150", "DELETED"])
    ws.append(["Widget D", "250", "ACTIVE"])
    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        drop_conditions=[
            {"column": "Product", "regex": "^TOTAL"},
            {"column": "Status", "equals": "DELETED"}
        ]
    )
    metas = loader.load_sheet(file_path, "multiple.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').df()

    assert len(result) == 3
    products = result["Product"].tolist()
    assert products == ["Widget A", "Widget B", "Widget D"]


def test_drop_conditions_column_not_found(temp_dir, loader):
    file_path = temp_dir / "missing_col.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Product"])
    ws.append(["Widget A"])
    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        drop_conditions=[
            {"column": "NonExistent", "regex": "test"}
        ]
    )
    metas = loader.load_sheet(file_path, "missing_col.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').df()

    assert len(result) == 1


def test_drop_conditions_with_drop_regex(temp_dir, loader):
    file_path = temp_dir / "both.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Product"])
    ws.append(["Widget A"])
    ws.append(["TOTAL"])
    ws.append(["Grand Total"])
    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        drop_regex="^Grand",
        drop_conditions=[
            {"column": "Product", "regex": "^TOTAL$"}
        ]
    )
    metas = loader.load_sheet(file_path, "both.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').df()

    assert len(result) == 1
    assert result["Product"].values[0] == "Widget A"


def test_drop_conditions_empty_dataframe(temp_dir, loader):
    file_path = temp_dir / "empty.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Name"])
    ws.append(["test1"])
    ws.append(["test2"])
    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        drop_conditions=[
            {"column": "Name", "regex": "test"}
        ]
    )
    metas = loader.load_sheet(file_path, "empty.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').df()

    assert len(result) == 0


def test_drop_conditions_all_rows_dropped(temp_dir, loader):
    file_path = temp_dir / "all_dropped.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Product"])
    ws.append(["Widget A"])
    ws.append(["Widget B"])
    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        drop_conditions=[
            {"column": "Product", "regex": ".*"}
        ]
    )
    metas = loader.load_sheet(file_path, "all_dropped.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').df()

    assert len(result) == 0


def test_drop_conditions_with_column_renames(temp_dir, loader):
    file_path = temp_dir / "rename.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ProductName"])
    ws.append(["Widget"])
    ws.append(["TOTAL"])
    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        column_renames={"ProductName": "Product"},
        drop_conditions=[
            {"column": "ProductName", "regex": "^TOTAL"}
        ]
    )
    metas = loader.load_sheet(file_path, "rename.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').df()

    assert len(result) == 1
    assert "Product" in result.columns or "product" in result.columns


def test_drop_conditions_string_equals(temp_dir, loader):
    file_path = temp_dir / "strings.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Name", "Category"])
    ws.append(["A", "Cat1"])
    ws.append(["B", "Cat2"])
    ws.append(["C", "Cat1"])
    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        drop_conditions=[
            {"column": "Category", "equals": "Cat1"}
        ]
    )
    metas = loader.load_sheet(file_path, "strings.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').df()

    assert len(result) == 1
    assert result["Name"].values[0] == "B"


def test_drop_conditions_multiple_columns_null(temp_dir, loader):
    file_path = temp_dir / "multi_null.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Name", "Value"])
    ws.append(["A", "100"])
    ws.append(["B", None])
    ws.append(["C", "300"])
    ws.append(["D", None])
    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        drop_conditions=[
            {"column": "Value", "is_null": True}
        ]
    )
    metas = loader.load_sheet(file_path, "multi_null.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').df()

    assert len(result) == 2
    names = result["Name"].tolist()
    assert "A" in names
    assert "C" in names


def test_drop_conditions_case_sensitive(temp_dir, loader):
    file_path = temp_dir / "case.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Product"])
    ws.append(["Total"])
    ws.append(["TOTAL"])
    ws.append(["total"])
    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        drop_conditions=[
            {"column": "Product", "regex": "^TOTAL$"}
        ]
    )
    metas = loader.load_sheet(file_path, "case.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').df()

    assert len(result) == 2
    products = result["Product"].tolist()
    assert "Total" in products
    assert "total" in products
    assert "TOTAL" not in products

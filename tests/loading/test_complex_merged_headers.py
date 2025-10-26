import pytest
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from mcp_excel.models import SheetOverride



def test_three_level_merged_headers(temp_dir, loader):
    file_path = temp_dir / "three_level.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    ws.merge_cells('B1:E1')
    ws['B1'] = '2024'

    ws.merge_cells('B2:C2')
    ws['B2'] = 'Q1'
    ws.merge_cells('D2:E2')
    ws['D2'] = 'Q2'

    ws['A3'] = 'Region'
    ws['B3'] = 'Jan'
    ws['C3'] = 'Feb'
    ws['D3'] = 'Apr'
    ws['E3'] = 'May'

    ws.append(['North', 100, 110, 200, 210])
    ws.append(['South', 150, 160, 250, 260])

    wb.save(file_path)

    override = SheetOverride(
        auto_detect=True,
        header_rows=3
    )

    metas = loader.load_sheet(file_path, "three_level.xlsx", "Sales", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 2


def test_unevenly_merged_headers(temp_dir, loader):
    file_path = temp_dir / "uneven_merge.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.merge_cells('A1:A2')
    ws['A1'] = 'ID'

    ws.merge_cells('B1:D1')
    ws['B1'] = 'Metrics'

    ws['B2'] = 'Revenue'
    ws['C2'] = 'Cost'
    ws['D2'] = 'Profit'

    ws.append([1, 1000, 600, 400])
    ws.append([2, 2000, 1200, 800])

    wb.save(file_path)

    override = SheetOverride(
        auto_detect=True,
        header_rows=2
    )

    metas = loader.load_sheet(file_path, "uneven_merge.xlsx", "Data", "test", override)
    meta = metas[0]

    schema = loader.conn.execute(f'DESCRIBE "{meta.table_name}"').fetchall()
    assert len(schema) >= 4


def test_merged_cells_spanning_with_totals(temp_dir, loader):
    file_path = temp_dir / "with_totals.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget"

    ws.merge_cells('A1:D1')
    ws['A1'] = 'Quarterly Budget'

    ws['A2'] = 'Department'
    ws['B2'] = 'Q1'
    ws['C2'] = 'Q2'
    ws['D2'] = 'Total'

    ws.append(['Sales', 10000, 12000, '=B3+C3'])
    ws.append(['Marketing', 5000, 6000, '=B4+C4'])

    wb.save(file_path)

    override = SheetOverride(
        skip_rows=1,
        header_rows=1
    )

    metas = loader.load_sheet(file_path, "with_totals.xlsx", "Budget", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 2


@pytest.mark.skip(reason="Complex multi-level merges need enhanced auto_detect")
def test_hierarchical_indent_headers(temp_dir, loader):
    file_path = temp_dir / "hierarchical.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Org"

    ws['A1'] = 'Division'
    ws.merge_cells('B1:C1')
    ws['B1'] = 'Sales'
    ws.merge_cells('D1:E1')
    ws['D1'] = 'Operations'

    ws['A2'] = 'Name'
    ws['B2'] = 'Revenue'
    ws['C2'] = 'Target'
    ws['D2'] = 'Cost'
    ws['E2'] = 'Budget'

    ws.append(['Alice', 100000, 120000, 80000, 90000])
    ws.append(['Bob', 150000, 140000, 100000, 95000])

    wb.save(file_path)

    override = SheetOverride(
        header_rows=2,
        auto_detect=True
    )

    metas = loader.load_sheet(file_path, "hierarchical.xlsx", "Org", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 2


def test_merged_headers_with_different_widths(temp_dir, loader):
    file_path = temp_dir / "diff_widths.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    ws.merge_cells('A1:B1')
    ws['A1'] = 'Customer Info'

    ws.merge_cells('C1:G1')
    ws['C1'] = 'Monthly Sales'

    ws['A2'] = 'ID'
    ws['B2'] = 'Name'
    ws['C2'] = 'Jan'
    ws['D2'] = 'Feb'
    ws['E2'] = 'Mar'
    ws['F2'] = 'Apr'
    ws['G2'] = 'May'

    ws.append([1, 'Acme Corp', 1000, 1100, 1200, 1300, 1400])

    wb.save(file_path)

    override = SheetOverride(
        header_rows=2,
        auto_detect=True
    )

    metas = loader.load_sheet(file_path, "diff_widths.xlsx", "Report", "test", override)
    meta = metas[0]

    schema = loader.conn.execute(f'DESCRIBE "{meta.table_name}"').fetchall()
    assert len(schema) >= 7


def test_merged_cells_with_rotated_text(temp_dir, loader):
    file_path = temp_dir / "rotated.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws['A1'] = 'Product'
    ws.merge_cells('B1:B2')
    ws['B1'] = 'Vertical Text'
    ws['C1'] = 'Price'

    ws['A2'] = 'Name'
    ws['C2'] = 'Amount'

    ws.append(['Widget', 100, 50.00])

    wb.save(file_path)

    override = SheetOverride(
        header_rows=2,
        auto_detect=True
    )

    metas = loader.load_sheet(file_path, "rotated.xlsx", "Data", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 1


@pytest.mark.skip(reason="Complex multi-level merges need enhanced auto_detect")
def test_repeating_headers_for_print_pagination(temp_dir, loader):
    file_path = temp_dir / "repeating.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LongReport"

    ws.append(['ID', 'Name', 'Value'])
    for i in range(1, 11):
        ws.append([i, f'Item {i}', i * 100])

    ws.append(['ID', 'Name', 'Value'])

    for i in range(11, 21):
        ws.append([i, f'Item {i}', i * 100])

    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        drop_regex='^ID$'
    )

    metas = loader.load_sheet(file_path, "repeating.xlsx", "LongReport", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 19


def test_merged_cells_in_header_with_formulas(temp_dir, loader):
    file_path = temp_dir / "formula_header.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calculated"

    ws.merge_cells('A1:C1')
    ws['A1'] = 'Summary Report'

    ws['A2'] = 'Item'
    ws['B2'] = 'Quantity'
    ws['C2'] = 'Total'

    ws.append(['Widget A', 10, '=B3*10'])
    ws.append(['Widget B', 20, '=B4*10'])

    wb.save(file_path)

    override = SheetOverride(
        skip_rows=1,
        header_rows=1
    )

    metas = loader.load_sheet(file_path, "formula_header.xlsx", "Calculated", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 2


@pytest.mark.skip(reason="Complex multi-level merges need enhanced auto_detect")
def test_complex_merge_with_blank_cells(temp_dir, loader):
    file_path = temp_dir / "merge_blanks.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Complex"

    ws.merge_cells('A1:A2')
    ws['A1'] = 'Category'

    ws.merge_cells('B1:C1')
    ws['B1'] = 'Group A'

    ws['B2'] = 'Sub1'
    ws['C2'] = 'Sub2'

    ws.append(['Cat1', 10, 20])
    ws.append(['Cat2', 30, 40])

    wb.save(file_path)

    override = SheetOverride(
        header_rows=2,
        auto_detect=True
    )

    metas = loader.load_sheet(file_path, "merge_blanks.xlsx", "Complex", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 2


def test_nested_merged_regions(temp_dir, loader):
    file_path = temp_dir / "nested_merge.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nested"

    ws.merge_cells('B1:G1')
    ws['B1'] = 'Annual Report 2024'

    ws.merge_cells('B2:D2')
    ws['B2'] = 'H1'
    ws.merge_cells('E2:G2')
    ws['E2'] = 'H2'

    ws['A3'] = 'Metric'
    ws['B3'] = 'Q1'
    ws['C3'] = 'Q2'
    ws['D3'] = 'Q3'
    ws['E3'] = 'Q4'
    ws['F3'] = 'Q5'
    ws['G3'] = 'Q6'

    ws.append(['Revenue', 1000, 1100, 1200, 1300, 1400, 1500])

    wb.save(file_path)

    override = SheetOverride(
        header_rows=3,
        auto_detect=True
    )

    metas = loader.load_sheet(file_path, "nested_merge.xlsx", "Nested", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 1

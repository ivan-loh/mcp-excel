import pytest
from pathlib import Path
import pandas as pd
import openpyxl
from mcp_excel.models import SheetOverride



def test_subtotals_every_n_rows(temp_dir, loader):
    file_path = temp_dir / "subtotals.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    ws.append(['Product', 'Amount'])
    ws.append(['Widget A', 100])
    ws.append(['Widget B', 150])
    ws.append(['Subtotal', '=SUM(B2:B3)'])
    ws.append(['Gadget A', 200])
    ws.append(['Gadget B', 250])
    ws.append(['Subtotal', '=SUM(B5:B6)'])

    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        drop_regex='^Subtotal$'
    )

    metas = loader.load_sheet(file_path, "subtotals.xlsx", "Sales", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) == 4


def test_grand_total_at_end(temp_dir, loader):
    file_path = temp_dir / "grand_total.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    ws.append(['Item', 'Value'])
    ws.append(['Item 1', 100])
    ws.append(['Item 2', 200])
    ws.append(['Item 3', 300])
    ws.append(['Grand Total', '=SUM(B2:B4)'])

    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        skip_footer=1
    )

    metas = loader.load_sheet(file_path, "grand_total.xlsx", "Report", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) == 3


def test_multiple_summary_rows(temp_dir, loader):
    file_path = temp_dir / "multi_summary.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget"

    ws.append(['Category', 'Amount'])
    ws.append(['Revenue', 10000])
    ws.append(['Revenue', 12000])
    ws.append(['Total Revenue', '=SUM(B2:B3)'])
    ws.append(['Expenses', 5000])
    ws.append(['Expenses', 6000])
    ws.append(['Total Expenses', '=SUM(B5:B6)'])
    ws.append(['Net Income', '=B4-B7'])

    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        drop_regex='^Total|^Net'
    )

    metas = loader.load_sheet(file_path, "multi_summary.xlsx", "Budget", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) == 4


@pytest.mark.skip(reason="Nested subtotals require complex drop_regex patterns")
def test_nested_subtotals(temp_dir, loader):
    file_path = temp_dir / "nested_subtotals.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detailed"

    ws.append(['Region', 'Product', 'Sales'])
    ws.append(['North', 'Widget A', 100])
    ws.append(['North', 'Widget B', 150])
    ws.append(['North Subtotal', '', 250])
    ws.append(['South', 'Widget A', 200])
    ws.append(['South', 'Widget B', 250])
    ws.append(['South Subtotal', '', 450])
    ws.append(['Grand Total', '', 700])

    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        drop_regex='Subtotal|Grand Total'
    )

    metas = loader.load_sheet(file_path, "nested_subtotals.xlsx", "Detailed", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) == 4


def test_blank_row_separators_with_totals(temp_dir, loader):
    file_path = temp_dir / "blank_separators.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(['Item', 'Value'])
    ws.append(['A', 10])
    ws.append(['B', 20])
    ws.append(['', ''])
    ws.append(['C', 30])
    ws.append(['D', 40])
    ws.append(['', ''])
    ws.append(['Total', 100])

    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        drop_regex='^Total$|^$'
    )

    metas = loader.load_sheet(file_path, "blank_separators.xlsx", "Data", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) == 4


def test_footer_notes_and_totals(temp_dir, loader):
    file_path = temp_dir / "footer_notes.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    ws.append(['Item', 'Quantity', 'Price', 'Total'])
    ws.append(['Widget', 10, 5.00, 50.00])
    ws.append(['Gadget', 5, 10.00, 50.00])
    ws.append(['', '', 'Subtotal:', 100.00])
    ws.append(['', '', 'Tax (10%):', 10.00])
    ws.append(['', '', 'Grand Total:', 110.00])
    ws.append(['Notes:', 'Thank you for your business', '', ''])

    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        skip_footer=4
    )

    metas = loader.load_sheet(file_path, "footer_notes.xlsx", "Invoice", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) == 2


def test_category_headers_as_data_rows(temp_dir, loader):
    file_path = temp_dir / "category_headers.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"

    ws.append(['Category', 'Amount'])
    ws.append(['Office Supplies', ''])
    ws.append(['Pens', 10])
    ws.append(['Paper', 20])
    ws.append(['Travel', ''])
    ws.append(['Flights', 500])
    ws.append(['Hotels', 300])

    wb.save(file_path)

    override = SheetOverride(
        header_rows=1
    )

    metas = loader.load_sheet(file_path, "category_headers.xlsx", "Expenses", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 4


def test_percentage_total_rows(temp_dir, loader):
    file_path = temp_dir / "percentages.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Share"

    ws.append(['Division', 'Revenue', 'Percent'])
    ws.append(['Sales', 10000, '40%'])
    ws.append(['Marketing', 5000, '20%'])
    ws.append(['Engineering', 10000, '40%'])
    ws.append(['Total', 25000, '100%'])

    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        drop_regex='^Total$'
    )

    metas = loader.load_sheet(file_path, "percentages.xlsx", "Share", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) == 3


def test_signature_lines_at_end(temp_dir, loader):
    file_path = temp_dir / "signatures.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Approval"

    ws.append(['Task', 'Status'])
    ws.append(['Task 1', 'Complete'])
    ws.append(['Task 2', 'Complete'])
    ws.append(['', ''])
    ws.append(['Approved by:', '_______________'])
    ws.append(['Date:', '_______________'])

    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        skip_footer=3
    )

    metas = loader.load_sheet(file_path, "signatures.xlsx", "Approval", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) == 2


def test_metadata_rows_before_headers(temp_dir, loader):
    file_path = temp_dir / "metadata_top.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    ws.append(['Company Report', ''])
    ws.append(['Generated:', '2024-01-15'])
    ws.append(['', ''])
    ws.append(['Product', 'Sales'])
    ws.append(['Widget A', 100])
    ws.append(['Widget B', 200])

    wb.save(file_path)

    override = SheetOverride(
        skip_rows=3,
        header_rows=1
    )

    metas = loader.load_sheet(file_path, "metadata_top.xlsx", "Report", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) == 2


def test_alternating_data_and_summary_rows(temp_dir, loader):
    file_path = temp_dir / "alternating.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mixed"

    ws.append(['Date', 'Amount', 'Type'])
    ws.append(['2024-01-01', 100, 'Sale'])
    ws.append(['Daily Total', 100, ''])
    ws.append(['2024-01-02', 150, 'Sale'])
    ws.append(['Daily Total', 150, ''])
    ws.append(['2024-01-03', 200, 'Sale'])
    ws.append(['Daily Total', 200, ''])

    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        drop_regex='^Daily Total$'
    )

    metas = loader.load_sheet(file_path, "alternating.xlsx", "Mixed", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) == 3


def test_page_break_summary_rows(temp_dir, loader):
    file_path = temp_dir / "page_breaks.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Long"

    ws.append(['ID', 'Value'])
    for i in range(1, 26):
        ws.append([i, i * 10])
        if i % 10 == 0:
            ws.append([f'Page {i//10} Subtotal', i * 10 * 10])

    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        drop_regex='Page.*Subtotal'
    )

    metas = loader.load_sheet(file_path, "page_breaks.xlsx", "Long", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 25

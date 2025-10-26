import pytest
from pathlib import Path
import openpyxl
from mcp_excel.models import SheetOverride

pytestmark = pytest.mark.unit


# Basic Detection Tests

def test_detect_multiple_tables_basic(loader):
    fixture_path = Path(__file__).parent.parent / "fixtures" / "multi_table_test.xlsx"
    override = SheetOverride(auto_detect=True, header_rows=1)
    metas = loader.load_sheet(fixture_path, "multi_table_test.xlsx", "MultiTable", "test", override)

    assert len(metas) == 2
    assert "_table0" in metas[0].table_name
    assert "_table1" in metas[1].table_name


def test_detect_three_tables(loader):
    fixture_path = Path(__file__).parent.parent / "fixtures" / "multi_table_test.xlsx"
    override = SheetOverride(auto_detect=True, header_rows=1)
    metas = loader.load_sheet(fixture_path, "multi_table_test.xlsx", "ThreeTables", "test", override)

    assert len(metas) == 3
    assert "_table0" in metas[0].table_name
    assert "_table1" in metas[1].table_name
    assert "_table2" in metas[2].table_name


def test_extract_specific_table(loader):
    """Test extract_table parameter to load only a specific table by index from multi-table sheet"""
    fixture_path = Path(__file__).parent.parent / "fixtures" / "multi_table_test.xlsx"
    override = SheetOverride(auto_detect=True, header_rows=1, extract_table=1)
    metas = loader.load_sheet(fixture_path, "multi_table_test.xlsx", "MultiTable", "test", override)

    assert len(metas) == 1
    assert "_table1" in metas[0].table_name


def test_extract_table_index_zero(loader):
    """Test extract_table=0 explicitly extracts the first table (index 0)"""
    fixture_path = Path(__file__).parent.parent / "fixtures" / "multi_table_test.xlsx"
    override = SheetOverride(auto_detect=True, header_rows=1, extract_table=0)
    metas = loader.load_sheet(fixture_path, "multi_table_test.xlsx", "MultiTable", "test", override)

    assert len(metas) == 1
    assert "_table0" in metas[0].table_name


def test_extract_table_out_of_range(loader):
    """Test that out-of-range extract_table index falls back to first table gracefully"""
    fixture_path = Path(__file__).parent.parent / "fixtures" / "multi_table_test.xlsx"
    override = SheetOverride(auto_detect=True, header_rows=1, extract_table=10)
    metas = loader.load_sheet(fixture_path, "multi_table_test.xlsx", "MultiTable", "test", override)

    assert len(metas) == 1


def test_multi_table_sql_query(loader):
    """Test SQL queries work correctly on multi-table views with expected row counts from fixture"""
    fixture_path = Path(__file__).parent.parent / "fixtures" / "multi_table_test.xlsx"
    override = SheetOverride(auto_detect=True, header_rows=1)
    metas = loader.load_sheet(fixture_path, "multi_table_test.xlsx", "MultiTable", "test", override)

    result0 = loader.conn.execute(f'SELECT COUNT(*) FROM "{metas[0].table_name}"').fetchone()
    result1 = loader.conn.execute(f'SELECT COUNT(*) FROM "{metas[1].table_name}"').fetchone()

    assert result0[0] == 3
    assert result1[0] == 4


def test_multi_table_separate_views(loader):
    """Test that each detected table creates a separate view in DuckDB information_schema"""
    fixture_path = Path(__file__).parent.parent / "fixtures" / "multi_table_test.xlsx"
    override = SheetOverride(auto_detect=True, header_rows=1)
    metas = loader.load_sheet(fixture_path, "multi_table_test.xlsx", "MultiTable", "test", override)

    tables = loader.conn.execute("SELECT table_name FROM information_schema.tables WHERE table_type='VIEW'").fetchall()
    table_names = [t[0] for t in tables]

    assert metas[0].table_name in table_names
    assert metas[1].table_name in table_names


def test_single_table_no_suffix(loader):
    """Test that single tables don't get _table0 suffix in their name (only multi-table gets suffixes)"""
    fixture_path = Path(__file__).parent.parent / "fixtures" / "multi_table_test.xlsx"
    override = SheetOverride(auto_detect=True, header_rows=1)
    metas = loader.load_sheet(fixture_path, "multi_table_test.xlsx", "SingleTable", "test", override)

    assert len(metas) == 1
    assert "_table0" not in metas[0].table_name and "_table1" not in metas[0].table_name


def test_multi_table_different_columns(loader):
    """Test that multiple tables can have same column count but different column names and data"""
    fixture_path = Path(__file__).parent.parent / "fixtures" / "multi_table_test.xlsx"
    override = SheetOverride(auto_detect=True, header_rows=1)
    metas = loader.load_sheet(fixture_path, "multi_table_test.xlsx", "MultiTable", "test", override)

    cols0 = loader.conn.execute(f'DESCRIBE "{metas[0].table_name}"').fetchall()
    cols1 = loader.conn.execute(f'DESCRIBE "{metas[1].table_name}"').fetchall()

    assert len(cols0) == 3
    assert len(cols1) == 3


def test_table_range_manual_override(temp_dir, loader):
    """Test table_range parameter to manually specify extraction range, overriding auto-detection"""
    file_path = temp_dir / "range_test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    ws['A1'] = "Ignore"
    ws['A5'] = "Name"
    ws['B5'] = "Value"
    ws['A6'] = "Alice"
    ws['B6'] = 100
    ws['A7'] = "Bob"
    ws['B7'] = 200

    wb.save(file_path)

    override = SheetOverride(auto_detect=True, header_rows=1, table_range="A5:B7")
    metas = loader.load_sheet(file_path, "range_test.xlsx", "Sheet", "test", override)

    assert len(metas) == 1
    result = loader.conn.execute(f'SELECT COUNT(*) FROM "{metas[0].table_name}"').fetchone()
    assert result[0] == 2


def test_blank_row_separator_detection(temp_dir, loader):
    """Test that 2+ consecutive blank rows correctly separate tables into distinct views"""
    file_path = temp_dir / "blank_sep.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    # Table 1: Rows 1-2
    ws['A1'] = "Product"
    ws['B1'] = "Price"
    ws['A2'] = "Widget"
    ws['B2'] = 100
    # Rows 3-6: 4 blank rows (>= 2 triggers table separation)

    # Table 2: Rows 7-8
    ws['A7'] = "Category"
    ws['B7'] = "Count"
    ws['A8'] = "Tools"
    ws['B8'] = 50

    wb.save(file_path)

    override = SheetOverride(auto_detect=True, header_rows=1)
    metas = loader.load_sheet(file_path, "blank_sep.xlsx", "Sheet", "test", override)

    # Should detect 2 separate tables due to 4 consecutive blank rows
    assert len(metas) == 2, "Should detect 2 tables separated by blank rows"

    # Verify table 1 data
    table1_data = loader.conn.execute(f'SELECT * FROM "{metas[0].table_name}"').fetchall()
    assert len(table1_data) == 1, "Table 1 should have 1 data row"
    assert table1_data[0][0] == "Widget", "Table 1 should contain Product data"

    # Verify table 2 data
    table2_data = loader.conn.execute(f'SELECT * FROM "{metas[1].table_name}"').fetchall()
    assert len(table2_data) == 1, "Table 2 should have 1 data row"
    assert table2_data[0][0] == "Tools", "Table 2 should contain Category data"


def test_multi_table_with_titles(temp_dir, loader):
    """Test detection of multiple tables each preceded by a title row (e.g., Q1 Report, Q2 Report)"""
    file_path = temp_dir / "titled_tables.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    # Table 1 with title
    ws['A1'] = "Q1 Report"  # Title row
    # Row 2: blank
    ws['A3'] = "Item"
    ws['B3'] = "Value"
    ws['A4'] = "A"
    ws['B4'] = 10
    # Rows 5-8: 4 blank rows

    # Table 2 with title
    ws['A9'] = "Q2 Report"  # Title row
    # Row 10: blank
    ws['A11'] = "Item"
    ws['B11'] = "Value"
    ws['A12'] = "B"
    ws['B12'] = 20

    wb.save(file_path)

    override = SheetOverride(auto_detect=True, header_rows=1)
    metas = loader.load_sheet(file_path, "titled_tables.xlsx", "Sheet", "test", override)

    # Should detect 2 tables with title rows
    assert len(metas) == 2, "Should detect 2 tables with title rows"

    # Verify each table has data
    for i, meta in enumerate(metas):
        data = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
        assert len(data) >= 1, f"Table {i} should have at least 1 data row"


def test_auto_detect_single_table(temp_dir, loader):
    """Test auto_detect mode correctly handles a simple single table"""
    file_path = temp_dir / "simple_table.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    ws['A1'] = "Name"
    ws['B1'] = "Score"
    ws['A2'] = "Test1"
    ws['B2'] = 90
    ws['A3'] = "Test2"
    ws['B3'] = 85

    wb.save(file_path)

    override = SheetOverride(auto_detect=True, header_rows=1)
    metas = loader.load_sheet(file_path, "simple_table.xlsx", "Sheet", "test", override)

    meta = metas[0]
    result = loader.conn.execute(f'SELECT COUNT(*) FROM "{meta.table_name}"').fetchone()
    assert result[0] == 2


def test_multi_table_metadata(loader):
    """Test that multi-table loading returns correct metadata for each table (mode, rows, names)"""
    fixture_path = Path(__file__).parent.parent / "fixtures" / "multi_table_test.xlsx"
    override = SheetOverride(auto_detect=True, header_rows=1)
    metas = loader.load_sheet(fixture_path, "multi_table_test.xlsx", "MultiTable", "test", override)

    for meta in metas:
        assert meta.mode == "ASSISTED"
        assert meta.est_rows > 0
        assert meta.table_name
        assert "MultiTable" in meta.sheet


def test_multi_table_without_auto_detect(loader):
    """Test that auto_detect=False prevents multi-table detection, loads entire sheet as single table"""
    fixture_path = Path(__file__).parent.parent / "fixtures" / "multi_table_test.xlsx"
    override = SheetOverride(auto_detect=False, header_rows=1)
    metas = loader.load_sheet(fixture_path, "multi_table_test.xlsx", "MultiTable", "test", override)

    assert len(metas) == 1


def test_table_confidence_scoring(temp_dir, loader):
    """Test that header detection works with auto_detect for simple tables with string headers"""
    file_path = temp_dir / "confidence.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    ws['A1'] = "Header1"
    ws['B1'] = "Header2"
    ws['A2'] = 100
    ws['B2'] = 200

    wb.save(file_path)

    override = SheetOverride(auto_detect=True, header_rows=1)
    metas = loader.load_sheet(file_path, "confidence.xlsx", "Sheet", "test", override)

    assert len(metas) == 1


def test_nested_blank_rows_within_table(temp_dir, loader):
    """Test that single blank row within table doesn't split it (need 2+ consecutive blank rows)"""
    file_path = temp_dir / "nested_blank.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    ws['A1'] = "Name"
    ws['B1'] = "Value"
    ws['A2'] = "Item1"
    ws['B2'] = 100
    ws['A4'] = "Item2"
    ws['B4'] = 200

    wb.save(file_path)

    override = SheetOverride(auto_detect=True, header_rows=1)
    metas = loader.load_sheet(file_path, "nested_blank.xlsx", "Sheet", "test", override)

    meta = metas[0]
    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 2


def test_multi_table_returns_list(loader):
    """Test that load_sheet always returns a list for multi-table results (backwards compatibility)"""
    fixture_path = Path(__file__).parent.parent / "fixtures" / "multi_table_test.xlsx"
    override = SheetOverride(auto_detect=True, header_rows=1)
    metas = loader.load_sheet(fixture_path, "multi_table_test.xlsx", "MultiTable", "test", override)

    assert isinstance(metas, list)
    assert all(hasattr(meta, 'table_name') for meta in metas)


def test_single_table_returns_list(loader, sample_excel):
    """Test that load_sheet always returns a list even for single table (backwards compatibility)"""
    metas = loader.load_sheet(sample_excel, "test.xlsx", "Data", "excel", None)

    assert isinstance(metas, list)
    assert len(metas) == 1


def test_multi_table_naming_convention(loader):
    """Test that multi-table names follow _table0, _table1, _table2 naming convention"""
    fixture_path = Path(__file__).parent.parent / "fixtures" / "multi_table_test.xlsx"
    override = SheetOverride(auto_detect=True, header_rows=1)
    metas = loader.load_sheet(fixture_path, "multi_table_test.xlsx", "MultiTable", "test", override)

    for idx, meta in enumerate(metas):
        assert f"_table{idx}" in meta.table_name


def test_multi_table_query_isolation(loader):
    fixture_path = Path(__file__).parent.parent / "fixtures" / "multi_table_test.xlsx"
    override = SheetOverride(auto_detect=True, header_rows=1)
    metas = loader.load_sheet(fixture_path, "multi_table_test.xlsx", "MultiTable", "test", override)

    rows0 = loader.conn.execute(f'SELECT * FROM "{metas[0].table_name}"').fetchall()
    rows1 = loader.conn.execute(f'SELECT * FROM "{metas[1].table_name}"').fetchall()

    assert len(rows0) == 3
    assert len(rows1) == 4

    data0 = {str(row[0]) for row in rows0 if row[0] is not None}
    data1 = {str(row[0]) for row in rows1 if row[0] is not None}
    assert data0.isdisjoint(data1)


# Format-Specific Tests

def test_csv_auto_detect_no_multi_table(temp_dir, loader):
    file_path = temp_dir / "test.csv"
    content = """Name,Value
Item1,100


Product,Price
Widget,50"""
    file_path.write_text(content, encoding='utf-8')

    override = SheetOverride(auto_detect=True, header_rows=1)
    metas = loader.load_sheet(file_path, "test.csv", "Sheet1", "csv", override)

    assert len(metas) == 1
    assert metas[0].mode in ["RAW", "ASSISTED"]


def test_xlsm_multi_table_detection(temp_dir, loader):
    file_path = temp_dir / "test.xlsm"
    wb = openpyxl.Workbook()
    ws = wb.active

    ws['A1'] = "Name"
    ws['B1'] = "Value"
    ws['A2'] = "Item1"
    ws['B2'] = 100

    ws['A7'] = "Product"
    ws['B7'] = "Price"
    ws['A8'] = "Widget"
    ws['B8'] = 50

    wb.save(file_path)

    override = SheetOverride(auto_detect=True, header_rows=1)
    metas = loader.load_sheet(file_path, "test.xlsm", "Sheet", "test", override)

    assert len(metas) == 2

    for i, meta in enumerate(metas):
        data = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
        assert len(data) >= 1

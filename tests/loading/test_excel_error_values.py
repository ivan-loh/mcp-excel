import pytest
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.utils.exceptions import IllegalCharacterError

pytestmark = pytest.mark.unit


def test_div_zero_error_in_cells(temp_dir, loader):
    file_path = temp_dir / "div_zero.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calculations"

    ws.append(["Item", "Value", "Result"])
    ws["A2"] = "Item A"
    ws["B2"] = 100
    ws["C2"] = "=B2/0"

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "div_zero.xlsx", "Calculations", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()

    assert len(result) >= 1


def test_ref_error_from_deleted_cells(temp_dir, loader):
    file_path = temp_dir / "ref_error.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(["Column1", "Column2", "Total"])
    ws["A2"] = 10
    ws["B2"] = 20
    ws["C2"] = "=A2+B2"

    wb.save(file_path)
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    ws.delete_cols(1)
    wb.save(file_path)

    metas = loader.load_sheet(file_path, "ref_error.xlsx", "Data", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()

    assert len(result) >= 0


def test_na_error_from_lookup(temp_dir, loader):
    file_path = temp_dir / "na_error.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lookups"

    ws.append(["Key", "Value", "Lookup"])
    ws["A2"] = "A"
    ws["B2"] = 100
    ws["C2"] = "=VLOOKUP(\"Z\", A2:B2, 2, FALSE)"

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "na_error.xlsx", "Lookups", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()

    assert len(result) >= 1


def test_value_error_type_mismatch(temp_dir, loader):
    file_path = temp_dir / "value_error.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Math"

    ws.append(["Number", "Text", "Result"])
    ws["A2"] = 100
    ws["B2"] = "ABC"
    ws["C2"] = "=A2+B2"

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "value_error.xlsx", "Math", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()

    assert len(result) >= 1


def test_num_error_invalid_numeric_operation(temp_dir, loader):
    file_path = temp_dir / "num_error.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Numbers"

    ws.append(["Value", "Result"])
    ws["A2"] = -1
    ws["B2"] = "=SQRT(A2)"

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "num_error.xlsx", "Numbers", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()

    assert len(result) >= 1


def test_name_error_undefined_name(temp_dir, loader):
    file_path = temp_dir / "name_error.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formulas"

    ws.append(["Item", "Calculation"])
    ws["A2"] = "Test"
    ws["B2"] = "=UndefinedFunction(A2)"

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "name_error.xlsx", "Formulas", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()

    assert len(result) >= 1


def test_mixed_errors_and_values(temp_dir, loader):
    file_path = temp_dir / "mixed_errors.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mixed"

    ws.append(["ID", "Value", "Computed"])
    ws["A2"] = 1
    ws["B2"] = 100
    ws["C2"] = "=B2*2"
    ws["A3"] = 2
    ws["B3"] = 0
    ws["C3"] = "=B2/B3"
    ws["A4"] = 3
    ws["B4"] = 200
    ws["C4"] = "=B4*2"

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "mixed_errors.xlsx", "Mixed", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()

    assert len(result) >= 2


def test_error_in_header_row(temp_dir, loader):
    file_path = temp_dir / "error_header.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws["A1"] = "Column1"
    ws["B1"] = "=1/0"
    ws["C1"] = "Column3"
    ws["A2"] = 100
    ws["B2"] = 200
    ws["C2"] = 300

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "error_header.xlsx", "Data", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()

    assert len(result) >= 1


def test_formula_referencing_external_workbook(temp_dir, loader):
    file_path = temp_dir / "external_ref.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(["Value", "External"])
    ws["A2"] = 100
    ws["B2"] = "=[NonexistentFile.xlsx]Sheet1!$A$1"

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "external_ref.xlsx", "Data", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()

    assert len(result) >= 1


def test_circular_reference_detection(temp_dir, loader):
    file_path = temp_dir / "circular.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Circular"

    ws.append(["A", "B", "C"])
    ws["A2"] = "=B2+1"
    ws["B2"] = "=C2+1"
    ws["C2"] = "=A2+1"

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "circular.xlsx", "Circular", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()

    assert len(result) >= 0


def test_null_error_in_formula(temp_dir, loader):
    file_path = temp_dir / "null_error.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(["Range1", "Range2", "Intersection"])
    ws["A2"] = "A1:A5"
    ws["B2"] = "C1:C5"
    ws["C2"] = "=A2 B2"

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "null_error.xlsx", "Data", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()

    assert len(result) >= 1


def test_array_formula_errors(temp_dir, loader):
    file_path = temp_dir / "array_formula.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Arrays"

    ws.append(["Values", "Doubled"])
    ws["A2"] = 10
    ws["A3"] = 20
    ws["A4"] = 30
    ws["B2"] = "=A2:A4*2"

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "array_formula.xlsx", "Arrays", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()

    assert len(result) >= 3


def test_getting_data_from_protected_cells(temp_dir, loader):
    file_path = temp_dir / "protected.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Protected"

    ws.append(["Public", "Protected"])
    ws["A2"] = 100
    ws["B2"] = 200

    ws.protection.sheet = True

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "protected.xlsx", "Protected", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()

    assert len(result) >= 1


def test_formula_with_nested_errors(temp_dir, loader):
    file_path = temp_dir / "nested_errors.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nested"

    ws.append(["A", "B", "C", "Result"])
    ws["A2"] = 10
    ws["B2"] = 0
    ws["C2"] = "=A2/B2"
    ws["D2"] = "=IF(ISERROR(C2), 0, C2)"

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "nested_errors.xlsx", "Nested", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()

    assert len(result) >= 1

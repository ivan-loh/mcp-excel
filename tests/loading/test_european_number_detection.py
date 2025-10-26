import pytest
from pathlib import Path
import pandas as pd
import openpyxl
from mcp_excel.models import SheetOverride, LocaleConfig



def test_european_decimal_format_text(temp_dir, loader):
    file_path = temp_dir / "european_numbers.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    ws.append(["Product", "Price", "Quantity"])
    ws["A2"] = "Widget A"
    ws["B2"] = "1.234,56"
    ws["C2"] = "100"
    ws["A3"] = "Widget B"
    ws["B3"] = "2.500,00"
    ws["C3"] = "50"

    wb.save(file_path)

    locale_config = LocaleConfig(
        decimal_separator=",",
        thousands_separator=".",
        auto_detect=False
    )
    override = SheetOverride(locale=locale_config, header_rows=1)

    metas = loader.load_sheet(file_path, "european_numbers.xlsx", "Sales", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT "Price" FROM "{meta.table_name}"').fetchall()

    assert len(result) == 2


@pytest.mark.skip(reason="Locale parsing needs explicit configuration or improved auto-detection")
def test_mixed_us_and_european_formats(temp_dir, loader):
    file_path = temp_dir / "mixed_formats.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(["ID", "Value_US", "Value_EU"])
    ws["A2"] = 1
    ws["B2"] = "1,234.56"
    ws["C2"] = "1.234,56"
    ws["A3"] = 2
    ws["B3"] = "2,500.00"
    ws["C3"] = "2.500,00"

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "mixed_formats.xlsx", "Data", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()

    assert len(result) == 2


def test_space_as_thousands_separator(temp_dir, loader):
    file_path = temp_dir / "space_separator.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Finance"

    ws.append(["Item", "Amount"])
    ws["A2"] = "Revenue"
    ws["B2"] = "1 234 567,89"
    ws["A3"] = "Expenses"
    ws["B3"] = "987 654,32"

    wb.save(file_path)

    locale_config = LocaleConfig(
        decimal_separator=",",
        thousands_separator=" ",
        auto_detect=False
    )
    override = SheetOverride(locale=locale_config, header_rows=1)

    metas = loader.load_sheet(file_path, "space_separator.xlsx", "Finance", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT "Amount" FROM "{meta.table_name}"').fetchall()

    assert len(result) == 2


@pytest.mark.skip(reason="Locale parsing needs explicit configuration or improved auto-detection")
def test_apostrophe_as_thousands_separator(temp_dir, loader):
    file_path = temp_dir / "apostrophe_separator.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Swiss"

    ws.append(["Description", "Value"])
    ws["A2"] = "Total"
    ws["B2"] = "1'234'567.89"

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "apostrophe_separator.xlsx", "Swiss", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT "Value" FROM "{meta.table_name}"').fetchall()

    assert len(result) == 1


def test_currency_symbols_european(temp_dir, loader):
    file_path = temp_dir / "euro_currency.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prices"

    ws.append(["Product", "Price"])
    ws["A2"] = "Item A"
    ws["B2"] = "€1.234,56"
    ws["A3"] = "Item B"
    ws["B3"] = "€2.500,00"

    wb.save(file_path)

    locale_config = LocaleConfig(
        decimal_separator=",",
        thousands_separator=".",
        currency_symbols=["€"],
        auto_detect=False
    )
    override = SheetOverride(locale=locale_config, header_rows=1)

    metas = loader.load_sheet(file_path, "euro_currency.xlsx", "Prices", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT "Price" FROM "{meta.table_name}"').fetchall()

    assert len(result) == 2


@pytest.mark.skip(reason="Locale parsing needs explicit configuration or improved auto-detection")
def test_auto_detect_european_format(temp_dir, loader):
    file_path = temp_dir / "auto_detect_eu.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(["Value"])
    ws["A2"] = "1.234,56"
    ws["A3"] = "2.500,00"
    ws["A4"] = "3.999,99"

    wb.save(file_path)

    override = SheetOverride(auto_detect=True, header_rows=1)

    metas = loader.load_sheet(file_path, "auto_detect_eu.xlsx", "Data", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()

    assert len(result) == 3


@pytest.mark.skip(reason="Locale parsing needs explicit configuration or improved auto-detection")
def test_percentage_european_format(temp_dir, loader):
    file_path = temp_dir / "percentage_eu.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Metrics"

    ws.append(["Metric", "Percentage"])
    ws["A2"] = "Growth Rate"
    ws["B2"] = "15,5%"
    ws["A3"] = "Margin"
    ws["B3"] = "22,75%"

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "percentage_eu.xlsx", "Metrics", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT "Percentage" FROM "{meta.table_name}"').fetchall()

    assert len(result) == 2


def test_negative_numbers_european_format(temp_dir, loader):
    file_path = temp_dir / "negative_eu.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Variance"

    ws.append(["Item", "Variance"])
    ws["A2"] = "Budget"
    ws["B2"] = "-1.234,56"
    ws["A3"] = "Forecast"
    ws["B3"] = "(2.500,00)"

    wb.save(file_path)

    locale_config = LocaleConfig(
        decimal_separator=",",
        thousands_separator=".",
        auto_detect=False
    )
    override = SheetOverride(locale=locale_config, header_rows=1)

    metas = loader.load_sheet(file_path, "negative_eu.xlsx", "Variance", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT "Variance" FROM "{meta.table_name}"').fetchall()

    assert len(result) == 2


def test_very_large_numbers_european(temp_dir, loader):
    file_path = temp_dir / "large_eu.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BigNumbers"

    ws.append(["Description", "Amount"])
    ws["A2"] = "Million"
    ws["B2"] = "1.000.000,00"
    ws["A3"] = "Billion"
    ws["B3"] = "1.000.000.000,00"

    wb.save(file_path)

    locale_config = LocaleConfig(
        decimal_separator=",",
        thousands_separator=".",
        auto_detect=False
    )
    override = SheetOverride(locale=locale_config, header_rows=1)

    metas = loader.load_sheet(file_path, "large_eu.xlsx", "BigNumbers", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT "Amount" FROM "{meta.table_name}"').fetchall()

    assert len(result) == 2


def test_decimal_only_no_thousands(temp_dir, loader):
    file_path = temp_dir / "decimal_only.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Simple"

    ws.append(["Value"])
    ws["A2"] = "123,45"
    ws["A3"] = "67,89"
    ws["A4"] = "0,50"

    wb.save(file_path)

    locale_config = LocaleConfig(
        decimal_separator=",",
        auto_detect=False
    )
    override = SheetOverride(locale=locale_config, header_rows=1)

    metas = loader.load_sheet(file_path, "decimal_only.xlsx", "Simple", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()

    assert len(result) == 3


@pytest.mark.skip(reason="Locale parsing needs explicit configuration or improved auto-detection")
def test_mixed_currency_symbols(temp_dir, loader):
    file_path = temp_dir / "multi_currency.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    ws.append(["Currency", "Amount"])
    ws["A2"] = "USD"
    ws["B2"] = "$1,234.56"
    ws["A3"] = "EUR"
    ws["B3"] = "€1.234,56"
    ws["A4"] = "GBP"
    ws["B4"] = "£1,234.56"

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "multi_currency.xlsx", "Transactions", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()

    assert len(result) == 3

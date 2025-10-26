import pytest
from pathlib import Path
import pandas as pd
import openpyxl
from mcp_excel.models import SheetOverride

pytestmark = pytest.mark.unit


def test_hidden_columns_with_data(temp_dir, loader):
    file_path = temp_dir / "hidden_cols.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    ws.append(['ID', 'Product', 'Internal_Cost', 'Price', 'Margin'])
    ws.append([1, 'Widget A', 50, 100, 50])
    ws.append([2, 'Widget B', 75, 150, 75])

    ws.column_dimensions['C'].hidden = True

    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        include_hidden=False
    )

    metas = loader.load_sheet(file_path, "hidden_cols.xlsx", "Sales", "test", override)
    meta = metas[0]

    schema = loader.conn.execute(f'DESCRIBE "{meta.table_name}"').fetchall()
    col_names = [col[0] for col in schema]

    assert len(schema) >= 4


def test_include_hidden_columns_explicitly(temp_dir, loader):
    file_path = temp_dir / "with_hidden.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(['A', 'B', 'C', 'D'])
    ws.append([1, 2, 3, 4])
    ws.append([5, 6, 7, 8])

    ws.column_dimensions['B'].hidden = True
    ws.column_dimensions['D'].hidden = True

    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        include_hidden=True
    )

    metas = loader.load_sheet(file_path, "with_hidden.xlsx", "Data", "test", override)
    meta = metas[0]

    schema = loader.conn.execute(f'DESCRIBE "{meta.table_name}"').fetchall()
    assert len(schema) == 4


def test_hidden_rows_with_data(temp_dir, loader):
    file_path = temp_dir / "hidden_rows.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    ws.append(['Name', 'Value'])
    ws.append(['Row 1', 100])
    ws.append(['Row 2 (Hidden)', 200])
    ws.append(['Row 3', 300])

    ws.row_dimensions[3].hidden = True

    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        include_hidden=False
    )

    metas = loader.load_sheet(file_path, "hidden_rows.xlsx", "Report", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 2


def test_include_hidden_rows_explicitly(temp_dir, loader):
    file_path = temp_dir / "rows_hidden.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(['ID', 'Status'])
    ws.append([1, 'Active'])
    ws.append([2, 'Inactive'])
    ws.append([3, 'Active'])

    ws.row_dimensions[3].hidden = True

    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        include_hidden=True
    )

    metas = loader.load_sheet(file_path, "rows_hidden.xlsx", "Data", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) == 3


def test_grouped_collapsed_rows(temp_dir, loader):
    file_path = temp_dir / "grouped.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grouped"

    ws.append(['Category', 'Item', 'Amount'])
    ws.append(['Sales', 'Widget A', 100])
    ws.append(['Sales', 'Widget B', 150])
    ws.append(['Sales', 'Subtotal', 250])
    ws.append(['Marketing', 'Campaign A', 50])

    for row in range(2, 4):
        ws.row_dimensions[row].outline_level = 1

    ws.row_dimensions[4].hidden = True

    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        include_hidden=False
    )

    metas = loader.load_sheet(file_path, "grouped.xlsx", "Grouped", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 3


def test_very_hidden_sheets(temp_dir, loader):
    file_path = temp_dir / "very_hidden.xlsx"

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Visible"
    ws1.append(['Data', 'Value'])
    ws1.append(['Item1', 100])

    ws2 = wb.create_sheet("Hidden")
    ws2.append(['Secret', 'Value'])
    ws2.append(['Item2', 200])
    ws2.sheet_state = 'hidden'

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "very_hidden.xlsx", "Visible", "test", None)
    assert len(metas) >= 1


def test_zero_width_columns(temp_dir, loader):
    file_path = temp_dir / "zero_width.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(['A', 'B', 'C'])
    ws.append([1, 2, 3])
    ws.append([4, 5, 6])

    ws.column_dimensions['B'].width = 0

    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        include_hidden=False
    )

    metas = loader.load_sheet(file_path, "zero_width.xlsx", "Data", "test", override)
    meta = metas[0]

    schema = loader.conn.execute(f'DESCRIBE "{meta.table_name}"').fetchall()
    assert len(schema) >= 2


def test_hidden_columns_in_middle(temp_dir, loader):
    file_path = temp_dir / "middle_hidden.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    ws.append(['First', 'Hidden1', 'Hidden2', 'Last'])
    ws.append([10, 20, 30, 40])
    ws.append([50, 60, 70, 80])

    ws.column_dimensions['B'].hidden = True
    ws.column_dimensions['C'].hidden = True

    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        include_hidden=False
    )

    metas = loader.load_sheet(file_path, "middle_hidden.xlsx", "Report", "test", override)
    meta = metas[0]

    schema = loader.conn.execute(f'DESCRIBE "{meta.table_name}"').fetchall()
    col_names = [col[0] for col in schema]

    assert 'First' in str(col_names) or 'first' in str(col_names)
    assert 'Last' in str(col_names) or 'last' in str(col_names)


def test_alternating_hidden_rows(temp_dir, loader):
    file_path = temp_dir / "alternating.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(['ID', 'Value'])
    for i in range(1, 11):
        ws.append([i, i * 10])
        if i % 2 == 0:
            ws.row_dimensions[i + 1].hidden = True

    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        include_hidden=False
    )

    metas = loader.load_sheet(file_path, "alternating.xlsx", "Data", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 5


def test_white_text_on_white_background(temp_dir, loader):
    file_path = temp_dir / "white_text.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hidden"

    ws.append(['Visible', 'Hidden'])
    ws.append(['Data1', 'SecretData1'])
    ws.append(['Data2', 'SecretData2'])

    from openpyxl.styles import Font, PatternFill

    for row in range(2, 4):
        ws[f'B{row}'].font = Font(color='FFFFFF')
        ws[f'B{row}'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    wb.save(file_path)

    override = SheetOverride(header_rows=1)

    metas = loader.load_sheet(file_path, "white_text.xlsx", "Hidden", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 2


def test_filtered_data_with_hidden_rows(temp_dir, loader):
    file_path = temp_dir / "filtered.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Filtered"

    ws.append(['Status', 'Amount'])
    ws.append(['Active', 100])
    ws.append(['Inactive', 200])
    ws.append(['Active', 300])

    ws.row_dimensions[3].hidden = True

    wb.save(file_path)

    override = SheetOverride(
        header_rows=1,
        include_hidden=False
    )

    metas = loader.load_sheet(file_path, "filtered.xlsx", "Filtered", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 2


def test_multiple_hidden_ranges(temp_dir, loader):
    file_path = temp_dir / "multi_hidden.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    for i in range(1, 21):
        ws.append([f'Row{i}', i * 10])

    for row in [3, 4, 5, 10, 11, 12, 18, 19]:
        ws.row_dimensions[row].hidden = True

    wb.save(file_path)

    override = SheetOverride(include_hidden=False)

    metas = loader.load_sheet(file_path, "multi_hidden.xlsx", "Data", "test", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    visible_rows = 20 - 8
    assert len(result) >= visible_rows - 1

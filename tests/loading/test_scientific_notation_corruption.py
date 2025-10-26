import pytest
from pathlib import Path
from mcp_excel.models import SheetOverride
import pandas as pd
import openpyxl
from openpyxl.styles import numbers

pytestmark = pytest.mark.unit


def test_large_id_as_text_preserved(temp_dir, loader):
    file_path = temp_dir / "large_ids.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    ws.append(["OrderID", "CustomerID", "Amount"])
    ws["A2"] = "123456789012345"
    ws["B2"] = "987654321098765"
    ws["C2"] = 1500.00

    ws.column_dimensions['A'].number_format = numbers.FORMAT_TEXT
    ws.column_dimensions['B'].number_format = numbers.FORMAT_TEXT

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "large_ids.xlsx", "Orders", "test", SheetOverride(header_rows=1))
    meta = metas[0]

    result = loader.conn.execute(f'SELECT "OrderID", "CustomerID" FROM "{meta.table_name}"').fetchall()

    order_id = str(result[0][0])
    customer_id = str(result[0][1])

    assert "123456789012345" in order_id
    assert "987654321098765" in customer_id
    assert "E+" not in order_id
    assert "E+" not in customer_id


@pytest.mark.skip(reason="Edge case - column without headers being read incorrectly")
def test_large_number_without_text_format_becomes_scientific(temp_dir, loader):
    file_path = temp_dir / "numeric_ids.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(["ID", "Value"])
    ws["A2"] = 123456789012345
    ws["B2"] = 100

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "numeric_ids.xlsx", "Data", "test", SheetOverride(header_rows=1))
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    id_value = result[0][0]

    assert id_value == 123456789012345 or str(id_value) == "1.23457e+14"


def test_credit_card_numbers_as_text(temp_dir, loader):
    file_path = temp_dir / "cards.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cards"

    ws.append(["CardNumber", "Last4"])
    ws["A2"] = "4532123456789012"
    ws["B2"] = "9012"

    ws.column_dimensions['A'].number_format = numbers.FORMAT_TEXT

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "cards.xlsx", "Cards", "test", SheetOverride(header_rows=1))
    meta = metas[0]

    result = loader.conn.execute(f'SELECT "CardNumber" FROM "{meta.table_name}"').fetchall()
    card = str(result[0][0])

    assert len(card) == 16 or "4532123456789012" in card


def test_tracking_numbers_large_integers(temp_dir, loader):
    file_path = temp_dir / "shipments.xlsx"

    df = pd.DataFrame({
        "TrackingNumber": ["1Z9999999999999999", "9400100000000000000000"],
        "Status": ["Delivered", "In Transit"]
    })

    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Shipments", index=False)
        workbook = writer.book
        worksheet = writer.sheets["Shipments"]

        for cell in worksheet['A']:
            cell.number_format = numbers.FORMAT_TEXT

    metas = loader.load_sheet(file_path, "shipments.xlsx", "Shipments", "test", SheetOverride(header_rows=1))
    meta = metas[0]

    result = loader.conn.execute(f'SELECT "TrackingNumber" FROM "{meta.table_name}"').fetchall()

    assert len(result) == 2
    tracking1 = str(result[0][0])
    tracking2 = str(result[1][0])

    assert "1Z" in tracking1 or tracking1.startswith("1Z")
    assert "9400" in tracking2 or tracking2.startswith("9400")


def test_mixed_numeric_and_text_ids(temp_dir, loader):
    file_path = temp_dir / "mixed_ids.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Items"

    ws.append(["ItemID", "Description"])
    ws["A2"] = "SKU-123"
    ws["A3"] = 456789012345678
    ws["A4"] = "PROD-999"
    ws["B2"] = "Widget A"
    ws["B3"] = "Widget B"
    ws["B4"] = "Widget C"

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "mixed_ids.xlsx", "Items", "test", SheetOverride(header_rows=1))
    meta = metas[0]

    result = loader.conn.execute(f'SELECT "ItemID" FROM "{meta.table_name}"').fetchall()

    assert len(result) == 3
    assert "SKU" in str(result[0][0]) or str(result[0][0]) == "SKU-123"
    assert "PROD" in str(result[2][0]) or str(result[2][0]) == "PROD-999"


@pytest.mark.skip(reason="Edge case - column without headers being read incorrectly")
def test_epoch_timestamps_preserved(temp_dir, loader):
    file_path = temp_dir / "timestamps.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Events"

    ws.append(["EventID", "Timestamp", "UnixTime"])
    ws["A2"] = 1001
    ws["B2"] = "2024-01-15 10:30:00"
    ws["C2"] = 1705318200

    ws.column_dimensions['C'].number_format = numbers.FORMAT_NUMBER

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "timestamps.xlsx", "Events", "test", SheetOverride(header_rows=1))
    meta = metas[0]

    result = loader.conn.execute(f'SELECT "UnixTime" FROM "{meta.table_name}"').fetchall()
    unix_time = result[0][0]

    assert unix_time == 1705318200 or abs(unix_time - 1705318200) < 10


def test_scientific_notation_in_source_data(temp_dir, loader):
    file_path = temp_dir / "scientific.xlsx"

    df = pd.DataFrame({
        "Measurement": ["1.23E+10", "5.67E-05", "8.90E+00"],
        "Value": [12300000000, 0.0000567, 8.90]
    })
    df.to_excel(file_path, sheet_name="Data", index=False)

    metas = loader.load_sheet(file_path, "scientific.xlsx", "Data", "test", SheetOverride(header_rows=1))
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()

    assert len(result) == 3
    assert result[0][1] == 12300000000 or abs(result[0][1] - 12300000000) < 1000

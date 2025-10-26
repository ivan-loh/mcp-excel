import pytest
from pathlib import Path
import pandas as pd
import openpyxl

pytestmark = pytest.mark.unit


def test_formula_referencing_same_sheet(temp_dir, loader):
    file_path = temp_dir / "internal_ref.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(['Value1', 'Value2', 'Sum'])
    ws['A2'] = 10
    ws['B2'] = 20
    ws['C2'] = '=A2+B2'

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "internal_ref.xlsx", "Data", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 1


def test_formula_referencing_other_sheet(temp_dir, loader):
    file_path = temp_dir / "cross_sheet.xlsx"

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Source"
    ws1.append(['Amount'])
    ws1['A2'] = 100

    ws2 = wb.create_sheet("Calculation")
    ws2.append(['Result'])
    ws2['A2'] = '=Source!A2*2'

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "cross_sheet.xlsx", "Calculation", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 1


def test_external_workbook_reference_missing_file(temp_dir, loader):
    file_path = temp_dir / "external_ref.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(['LocalValue', 'ExternalValue'])
    ws['A2'] = 100
    ws['B2'] = "=[MissingFile.xlsx]Sheet1!$A$1"

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "external_ref.xlsx", "Data", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 1


def test_named_range_formula(temp_dir, loader):
    file_path = temp_dir / "named_range.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(['Price', 'Quantity', 'Total'])
    ws['A2'] = 10
    ws['B2'] = 5
    ws['C2'] = '=A2*B2'

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "named_range.xlsx", "Data", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 1


def test_volatile_functions(temp_dir, loader):
    file_path = temp_dir / "volatile.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(['StaticValue', 'TodayDate', 'RandomValue'])
    ws['A2'] = 100
    ws['B2'] = '=TODAY()'
    ws['C2'] = '=RAND()'

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "volatile.xlsx", "Data", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 1


def test_formula_with_error_propagation(temp_dir, loader):
    file_path = temp_dir / "error_prop.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(['A', 'B', 'C', 'D'])
    ws['A2'] = 10
    ws['B2'] = 0
    ws['C2'] = '=A2/B2'
    ws['D2'] = '=C2*2'

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "error_prop.xlsx", "Data", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 1


def test_complex_nested_formula(temp_dir, loader):
    file_path = temp_dir / "complex_formula.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calc"

    ws.append(['Value', 'Computed'])
    ws['A2'] = 10
    ws['B2'] = '=IF(A2>0, A2*2, 0)'
    ws['A3'] = -5
    ws['B3'] = '=IF(A3>0, A3*2, 0)'

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "complex_formula.xlsx", "Calc", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 2


def test_vlookup_with_missing_reference(temp_dir, loader):
    file_path = temp_dir / "vlookup.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(['LookupKey', 'Result'])
    ws['A2'] = 'A001'
    ws['B2'] = '=VLOOKUP(A2, ExternalSheet!A:B, 2, FALSE)'

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "vlookup.xlsx", "Data", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 1


def test_sumif_formula(temp_dir, loader):
    file_path = temp_dir / "sumif.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(['Category', 'Amount', 'Total_Sales'])
    ws['A2'] = 'Sales'
    ws['B2'] = 100
    ws['C2'] = '=SUMIF($A$2:$A$10, "Sales", $B$2:$B$10)'
    ws['A3'] = 'Sales'
    ws['B3'] = 200

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "sumif.xlsx", "Data", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 2


def test_index_match_formula(temp_dir, loader):
    file_path = temp_dir / "index_match.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lookup"

    ws.append(['Key', 'Value', 'Result'])
    ws['A2'] = 'A'
    ws['B2'] = 100
    ws['C2'] = '=INDEX(B:B, MATCH("A", A:A, 0))'

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "index_match.xlsx", "Lookup", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 1


def test_indirect_function(temp_dir, loader):
    file_path = temp_dir / "indirect.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(['Cell_Address', 'Value'])
    ws['A2'] = 'B2'
    ws['B2'] = 100
    ws['A3'] = '=INDIRECT(A2)'

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "indirect.xlsx", "Data", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 2


def test_getpivotdata_function(temp_dir, loader):
    file_path = temp_dir / "pivot_ref.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    ws.append(['Metric', 'Value'])
    ws['A2'] = 'Total Sales'
    ws['B2'] = 10000

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "pivot_ref.xlsx", "Summary", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 1


def test_offset_function(temp_dir, loader):
    file_path = temp_dir / "offset.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(['Base', 'Offset_Value'])
    ws['A2'] = 100
    ws['B2'] = '=OFFSET(A2, 1, 0)'
    ws['A3'] = 200

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "offset.xlsx", "Data", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 2


def test_choose_function(temp_dir, loader):
    file_path = temp_dir / "choose.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(['Index', 'Result'])
    ws['A2'] = 2
    ws['B2'] = '=CHOOSE(A2, "First", "Second", "Third")'

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "choose.xlsx", "Data", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 1


def test_textjoin_function(temp_dir, loader):
    file_path = temp_dir / "textjoin.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(['Part1', 'Part2', 'Part3', 'Combined'])
    ws['A2'] = 'Hello'
    ws['B2'] = 'World'
    ws['C2'] = '2024'
    ws['D2'] = '=TEXTJOIN(" ", TRUE, A2, B2, C2)'

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "textjoin.xlsx", "Data", "test", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 1

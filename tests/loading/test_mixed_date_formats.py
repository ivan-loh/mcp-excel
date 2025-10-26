import pytest
from pathlib import Path
from mcp_excel.models import SheetOverride
import pandas as pd
import openpyxl
from datetime import datetime
from openpyxl.styles import numbers

pytestmark = pytest.mark.unit


def test_mixed_date_formats_in_column(temp_dir, loader):
    file_path = temp_dir / "mixed_dates.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    ws.append(["ID", "Date", "Amount"])
    ws["A2"] = 1
    ws["B2"] = "2024-01-15"
    ws["C2"] = 100
    ws["A3"] = 2
    ws["B3"] = "01/15/2024"
    ws["C3"] = 200
    ws["A4"] = 3
    ws["B4"] = "Jan 15, 2024"
    ws["C4"] = 300

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "mixed_dates.xlsx", "Transactions", "test", SheetOverride(header_rows=1))
    meta = metas[0]

    result = loader.conn.execute(f'SELECT "Date" FROM "{meta.table_name}"').fetchall()

    assert len(result) == 3


def test_date_as_excel_serial_number(temp_dir, loader):
    file_path = temp_dir / "serial_dates.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(["OrderID", "OrderDate", "Amount"])
    ws["A2"] = 1001
    ws["B2"] = datetime(2024, 1, 15)
    ws["C2"] = 500.00

    ws["B2"].number_format = numbers.FORMAT_DATE_XLSX14

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "serial_dates.xlsx", "Data", "test", SheetOverride(header_rows=1))
    meta = metas[0]

    result = loader.conn.execute(f'SELECT "OrderDate" FROM "{meta.table_name}"').fetchall()

    assert len(result) == 1
    date_val = result[0][0]

    assert date_val is not None


def test_text_dates_not_recognized_by_excel(temp_dir, loader):
    file_path = temp_dir / "text_dates.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Events"

    ws.append(["EventID", "EventDate", "Description"])
    ws["A2"] = 1
    ws["B2"] = "15-Jan-2024"
    ws["C2"] = "Conference"
    ws["A3"] = 2
    ws["B3"] = "2024/01/15"
    ws["C3"] = "Meeting"
    ws["A4"] = 3
    ws["B4"] = "15.01.2024"
    ws["C4"] = "Workshop"

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "text_dates.xlsx", "Events", "test", SheetOverride(header_rows=1))
    meta = metas[0]

    result = loader.conn.execute(f'SELECT "EventDate" FROM "{meta.table_name}"').fetchall()

    assert len(result) == 3


def test_ambiguous_dates_us_vs_european(temp_dir, loader):
    file_path = temp_dir / "ambiguous_dates.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shipments"

    ws.append(["ShipmentID", "Date", "Status"])
    ws["A2"] = 1
    ws["B2"] = "01/02/2024"
    ws["C2"] = "Delivered"
    ws["A3"] = 2
    ws["B3"] = "02/01/2024"
    ws["C3"] = "In Transit"
    ws["A4"] = 3
    ws["B4"] = "13/01/2024"
    ws["C4"] = "Delivered"

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "ambiguous_dates.xlsx", "Shipments", "test", SheetOverride(header_rows=1))
    meta = metas[0]

    result = loader.conn.execute(f'SELECT "Date" FROM "{meta.table_name}"').fetchall()

    assert len(result) == 3


def test_dates_with_times_mixed_formats(temp_dir, loader):
    file_path = temp_dir / "datetime_mixed.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Logs"

    ws.append(["LogID", "Timestamp", "Message"])
    ws["A2"] = 1
    ws["B2"] = "2024-01-15 10:30:00"
    ws["C2"] = "Event A"
    ws["A3"] = 2
    ws["B3"] = "01/15/2024 14:45"
    ws["C3"] = "Event B"
    ws["A4"] = 3
    ws["B4"] = datetime(2024, 1, 15, 18, 20, 0)
    ws["C4"] = "Event C"

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "datetime_mixed.xlsx", "Logs", "test", SheetOverride(header_rows=1))
    meta = metas[0]

    result = loader.conn.execute(f'SELECT "Timestamp" FROM "{meta.table_name}"').fetchall()

    assert len(result) == 3


def test_null_dates_different_representations(temp_dir, loader):
    file_path = temp_dir / "null_dates.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projects"

    ws.append(["ProjectID", "StartDate", "EndDate"])
    ws["A2"] = 1
    ws["B2"] = datetime(2024, 1, 1)
    ws["C2"] = datetime(2024, 6, 30)
    ws["A3"] = 2
    ws["B3"] = datetime(2024, 2, 1)
    ws["C3"] = "TBD"
    ws["A4"] = 3
    ws["B4"] = datetime(2024, 3, 1)
    ws["C4"] = None

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "null_dates.xlsx", "Projects", "test", SheetOverride(header_rows=1))
    meta = metas[0]

    result = loader.conn.execute(f'SELECT "EndDate" FROM "{meta.table_name}"').fetchall()

    assert len(result) == 3


def test_fiscal_quarter_text_dates(temp_dir, loader):
    file_path = temp_dir / "fiscal_dates.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financials"

    ws.append(["Period", "Revenue", "Quarter"])
    ws["A2"] = "Q1 2024"
    ws["B2"] = 100000
    ws["C2"] = "Q1"
    ws["A3"] = "Q2 2024"
    ws["B3"] = 120000
    ws["C3"] = "Q2"
    ws["A4"] = "2024-Q3"
    ws["B4"] = 130000
    ws["C4"] = "Q3"

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "fiscal_dates.xlsx", "Financials", "test", SheetOverride(header_rows=1))
    meta = metas[0]

    result = loader.conn.execute(f'SELECT "Period", "Quarter" FROM "{meta.table_name}"').fetchall()

    assert len(result) == 3


def test_relative_date_strings(temp_dir, loader):
    file_path = temp_dir / "relative_dates.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"

    ws.append(["TaskID", "DueDate", "Status"])
    ws["A2"] = 1
    ws["B2"] = "Today"
    ws["C2"] = "Active"
    ws["A3"] = 2
    ws["B3"] = "Tomorrow"
    ws["C3"] = "Pending"
    ws["A4"] = 3
    ws["B4"] = datetime(2024, 1, 20)
    ws["C4"] = "Active"

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "relative_dates.xlsx", "Tasks", "test", SheetOverride(header_rows=1))
    meta = metas[0]

    result = loader.conn.execute(f'SELECT "DueDate" FROM "{meta.table_name}"').fetchall()

    assert len(result) == 3


def test_year_only_dates(temp_dir, loader):
    file_path = temp_dir / "year_dates.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "History"

    ws.append(["EventID", "Year", "Description"])
    ws["A2"] = 1
    ws["B2"] = 2020
    ws["C2"] = "Event A"
    ws["A3"] = 2
    ws["B3"] = "2021"
    ws["C3"] = "Event B"
    ws["A4"] = 3
    ws["B4"] = 2022
    ws["C4"] = "Event C"

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "year_dates.xlsx", "History", "test", SheetOverride(header_rows=1))
    meta = metas[0]

    result = loader.conn.execute(f'SELECT "Year" FROM "{meta.table_name}"').fetchall()

    assert len(result) == 3


def test_month_year_various_formats(temp_dir, loader):
    file_path = temp_dir / "month_year.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly"

    ws.append(["Period", "Sales"])
    ws["A2"] = "Jan 2024"
    ws["B2"] = 10000
    ws["A3"] = "2024-02"
    ws["B3"] = 12000
    ws["A4"] = "March 2024"
    ws["B4"] = 11000

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "month_year.xlsx", "Monthly", "test", SheetOverride(header_rows=1))
    meta = metas[0]

    result = loader.conn.execute(f'SELECT "Period" FROM "{meta.table_name}"').fetchall()

    assert len(result) == 3


def test_iso_8601_dates(temp_dir, loader):
    file_path = temp_dir / "iso_dates.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "API_Logs"

    ws.append(["RequestID", "Timestamp", "Status"])
    ws["A2"] = 1
    ws["B2"] = "2024-01-15T10:30:00Z"
    ws["C2"] = 200
    ws["A3"] = 2
    ws["B3"] = "2024-01-15T14:45:00+00:00"
    ws["C3"] = 200
    ws["A4"] = 3
    ws["B4"] = datetime(2024, 1, 15, 18, 20, 0)
    ws["C4"] = 200

    wb.save(file_path)

    metas = loader.load_sheet(file_path, "iso_dates.xlsx", "API_Logs", "test", SheetOverride(header_rows=1))
    meta = metas[0]

    result = loader.conn.execute(f'SELECT "Timestamp" FROM "{meta.table_name}"').fetchall()

    assert len(result) == 3

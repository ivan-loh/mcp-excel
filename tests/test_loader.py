import pytest
from pathlib import Path
import pandas as pd
import openpyxl
from mcp_excel.types import SheetOverride, MergeHandlingConfig

pytestmark = pytest.mark.unit


def test_load_raw_mode(loader, sample_excel):
    metas = loader.load_sheet(sample_excel, "test.xlsx", "Data", "excel", None)
    meta = metas[0]
    assert meta.mode == "RAW"
    assert meta.table_name
    assert meta.est_rows > 0


def test_load_assisted_mode_basic(loader, sample_excel):
    override = SheetOverride(header_rows=1)
    metas = loader.load_sheet(sample_excel, "test.xlsx", "Data", "excel", override)
    meta = metas[0]
    assert meta.mode == "ASSISTED"
    assert meta.est_rows > 0


def test_skip_rows(temp_dir, loader):
    file_path = temp_dir / "skip_test.xlsx"
    df = pd.DataFrame({
        "Header": ["Skip", "Skip", "Name", "Alice", "Bob"],
        "Col2": ["Skip", "Skip", "Age", "25", "30"]
    })
    df.to_excel(file_path, sheet_name="Data", index=False, header=False)

    override = SheetOverride(skip_rows=2, header_rows=1)
    metas = loader.load_sheet(file_path, "skip_test.xlsx", "Data", "excel", override)
    meta = metas[0]
    assert meta.mode == "ASSISTED"


def test_skip_footer(temp_dir, loader):
    file_path = temp_dir / "footer_test.xlsx"
    df = pd.DataFrame({
        "Name": ["Alice", "Bob", "Total", "Notes"],
        "Age": ["25", "30", "55", "End"]
    })
    df.to_excel(file_path, sheet_name="Data", index=False)

    override = SheetOverride(skip_footer=2, header_rows=1)
    metas = loader.load_sheet(file_path, "footer_test.xlsx", "Data", "excel", override)
    meta = metas[0]
    assert meta.est_rows == 2


def test_drop_regex(temp_dir, loader):
    file_path = temp_dir / "regex_test.xlsx"
    df = pd.DataFrame({
        "Name": ["Alice", "Bob", "Total:", "Notes:"],
        "Value": [100, 200, 300, 0]
    })
    df.to_excel(file_path, sheet_name="Data", index=False)

    override = SheetOverride(drop_regex="^(Total|Notes):", header_rows=1)
    metas = loader.load_sheet(file_path, "regex_test.xlsx", "Data", "excel", override)
    meta = metas[0]
    assert meta.est_rows == 2


def test_column_renames(temp_dir, loader):
    file_path = temp_dir / "rename_test.xlsx"
    df = pd.DataFrame({
        "OldName": ["A", "B"],
        "AnotherOld": [1, 2]
    })
    df.to_excel(file_path, sheet_name="Data", index=False)

    override = SheetOverride(column_renames={"OldName": "NewName"}, header_rows=1)
    metas = loader.load_sheet(file_path, "rename_test.xlsx", "Data", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'DESCRIBE "{meta.table_name}"').fetchall()
    column_names = [row[0] for row in result]
    assert "NewName" in column_names or "newname" in column_names


def test_get_sheet_names(loader, sample_excel):
    sheets = loader.get_sheet_names(sample_excel)
    assert "Data" in sheets
    assert len(sheets) >= 1


def test_multirow_headers(temp_dir, loader):
    file_path = temp_dir / "multirow_test.xlsx"
    data = [
        ["Region", "Q1", "Q1", "Q2", "Q2"],
        ["", "Sales", "Units", "Sales", "Units"],
        ["North", 1000, 50, 1200, 60],
        ["South", 800, 40, 900, 45]
    ]
    df = pd.DataFrame(data)
    df.to_excel(file_path, sheet_name="Data", index=False, header=False)

    override = SheetOverride(header_rows=2)
    metas = loader.load_sheet(file_path, "multirow_test.xlsx", "Data", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}" LIMIT 1').fetchall()
    assert len(result) > 0
    assert meta.est_rows == 2


def test_type_hints(temp_dir, loader):
    file_path = temp_dir / "types_test.xlsx"
    df = pd.DataFrame({
        "Name": ["Alice", "Bob"],
        "Age": ["25", "30"],
        "Salary": ["50000.50", "60000.75"],
        "Active": ["true", "false"]
    })
    df.to_excel(file_path, sheet_name="Data", index=False)

    override = SheetOverride(
        header_rows=1,
        type_hints={
            "Age": "INT",
            "Salary": "DECIMAL",
            "Active": "BOOL"
        }
    )
    metas = loader.load_sheet(file_path, "types_test.xlsx", "Data", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}" LIMIT 1').fetchall()
    assert len(result) > 0


def test_unpivot(temp_dir, loader):
    file_path = temp_dir / "unpivot_test.xlsx"
    df = pd.DataFrame({
        "Region": ["North", "South"],
        "Jan": [100, 80],
        "Feb": [110, 85],
        "Mar": [120, 90]
    })
    df.to_excel(file_path, sheet_name="Data", index=False)

    override = SheetOverride(
        header_rows=1,
        unpivot={
            "id_vars": ["Region"],
            "value_vars": ["Jan", "Feb", "Mar"],
            "var_name": "Month",
            "value_name": "Sales"
        }
    )
    metas = loader.load_sheet(file_path, "unpivot_test.xlsx", "Data", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT COUNT(*) FROM "{meta.table_name}"').fetchone()
    assert result[0] == 6


def test_merged_cells_horizontal_headers(temp_dir, loader):
    file_path = temp_dir / "merged_headers.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(['Name', 'Value'])
    ws.append(['Item1', 100])
    ws.append(['Item2', 200])

    ws.merge_cells('A1:B1')
    ws['A1'] = 'Header'

    wb.save(file_path)

    override = SheetOverride(
        auto_detect=True,
        merge_handling=MergeHandlingConfig(strategy='fill'),
        header_rows=1
    )
    metas = loader.load_sheet(file_path, "merged_headers.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) == 2
    assert meta.est_rows == 2


def test_merged_cells_auto_detect(temp_dir, loader):
    file_path = temp_dir / "merged_auto.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.merge_cells('A1:B1')
    ws['A1'] = 'Region'

    ws.append(['Name', 'Value'])
    ws.append(['North', 100])
    ws.append(['South', 200])

    wb.save(file_path)

    override = SheetOverride(auto_detect=True, header_rows=1)
    metas = loader.load_sheet(file_path, "merged_auto.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    assert meta.est_rows == 2


def test_merged_cells_data_area(temp_dir, loader):
    file_path = temp_dir / "merged_data.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(['Name', 'Value'])
    ws.append(['Item1', 100])
    ws.merge_cells('A3:B3')
    ws['A3'] = 'Merged Data'
    ws.append(['Item2', 200])

    wb.save(file_path)

    override = SheetOverride(
        auto_detect=True,
        merge_handling=MergeHandlingConfig(strategy='fill'),
        header_rows=1
    )
    metas = loader.load_sheet(file_path, "merged_data.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) >= 3


def test_hidden_rows_ignored(temp_dir, loader):
    file_path = temp_dir / "hidden_rows.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(['Report Generated:', '2024-01-01'])
    ws.append(['Department:', 'Sales'])
    ws.row_dimensions[1].hidden = True
    ws.row_dimensions[2].hidden = True

    ws.append(['Name', 'Value'])
    ws.append(['Alice', 100])
    ws.append(['Bob', 200])

    ws.append(['TOTAL', 300])
    ws.row_dimensions[6].hidden = True

    wb.save(file_path)

    override = SheetOverride(auto_detect=True, include_hidden=False, header_rows=1)
    metas = loader.load_sheet(file_path, "hidden_rows.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) == 2
    assert meta.est_rows == 2


def test_hidden_rows_included(temp_dir, loader):
    file_path = temp_dir / "hidden_rows_include.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(['Name', 'Value'])
    ws.append(['Alice', 100])
    ws.row_dimensions[2].hidden = True
    ws.append(['Bob', 200])

    wb.save(file_path)

    override = SheetOverride(auto_detect=True, include_hidden=True, header_rows=1)
    metas = loader.load_sheet(file_path, "hidden_rows_include.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) == 2


def test_hidden_columns_ignored(temp_dir, loader):
    file_path = temp_dir / "hidden_cols.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(['Name', 'Hidden', 'Value'])
    ws.append(['Alice', 'X', 100])
    ws.append(['Bob', 'Y', 200])
    ws.column_dimensions['B'].hidden = True

    wb.save(file_path)

    override = SheetOverride(auto_detect=True, include_hidden=False, header_rows=1)
    metas = loader.load_sheet(file_path, "hidden_cols.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'DESCRIBE "{meta.table_name}"').fetchall()
    column_names = [row[0].lower() for row in result]
    assert 'hidden' not in column_names


def test_european_number_format_csv(temp_dir, loader):
    file_path = temp_dir / "european.csv"
    content = """Name;Amount;Date
Product A;1.234,56;15.01.2024
Product B;2.500,00;20.01.2024
Product C;999,99;25.01.2024"""

    file_path.write_text(content, encoding='utf-8')

    override = SheetOverride(auto_detect=True, header_rows=1)
    metas = loader.load_sheet(file_path, "european.csv", "Sheet1", "csv", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}" LIMIT 1').fetchall()
    assert len(result) == 1


def test_csv_utf8_with_bom(temp_dir, loader):
    file_path = temp_dir / "utf8_bom.csv"
    content = "Name,Value\nCafé,100\nNaïve,200"
    file_path.write_text(content, encoding='utf-8-sig')

    metas = loader.load_sheet(file_path, "utf8_bom.csv", "Sheet1", "csv", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) == 2


def test_csv_latin1_encoding(temp_dir, loader):
    file_path = temp_dir / "latin1.csv"
    content = "Name,Value\nCafé,100\nNaïve,200\nZürich,300"
    file_path.write_text(content, encoding='latin-1')

    metas = loader.load_sheet(file_path, "latin1.csv", "Sheet1", "csv", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) == 3


def test_csv_windows1252_encoding(temp_dir, loader):
    file_path = temp_dir / "windows.csv"
    content = "Name,Value\nTest™,100\nData©,200"
    file_path.write_text(content, encoding='windows-1252')

    metas = loader.load_sheet(file_path, "windows.csv", "Sheet1", "csv", None)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) == 2


def test_hidden_metadata_rows(temp_dir, loader):
    file_path = temp_dir / "metadata.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(['Report Title:', 'Sales Report'])
    ws.append(['Generated:', '2024-01-01'])
    ws.append(['Department:', 'Sales'])
    ws.row_dimensions[1].hidden = True
    ws.row_dimensions[2].hidden = True
    ws.row_dimensions[3].hidden = True

    ws.append(['Product', 'Revenue'])
    ws.append(['Widget', 1000])
    ws.append(['Gadget', 2000])

    wb.save(file_path)

    override = SheetOverride(auto_detect=True, include_hidden=False, header_rows=1)
    metas = loader.load_sheet(file_path, "metadata.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) == 2
    assert result[0][0] == 'Widget'


def test_hidden_summary_rows(temp_dir, loader):
    file_path = temp_dir / "summary.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(['Product', 'Amount'])
    ws.append(['Item1', 100])
    ws.append(['Item2', 200])
    ws.append(['Item3', 300])
    ws.append(['TOTAL', 600])
    ws.append(['AVERAGE', 200])
    ws.row_dimensions[5].hidden = True
    ws.row_dimensions[6].hidden = True

    wb.save(file_path)

    override = SheetOverride(auto_detect=True, include_hidden=False, header_rows=1)
    metas = loader.load_sheet(file_path, "summary.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) == 3
    assert all(row[0] in ['Item1', 'Item2', 'Item3'] for row in result)


def test_number_format_locale_detection(temp_dir, loader):
    file_path = temp_dir / "locale_detect.csv"
    content = """Product;Price;Quantity
Widget;1.234,56;100
Gadget;2.500,00;200
Tool;999,99;150"""
    file_path.write_text(content, encoding='utf-8')

    override = SheetOverride(auto_detect=True, header_rows=1)
    metas = loader.load_sheet(file_path, "locale_detect.csv", "Sheet1", "csv", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}" LIMIT 1').fetchall()
    assert len(result) == 1


def test_csv_encoding_override(temp_dir, loader):
    file_path = temp_dir / "override_encoding.csv"
    content = "Name,Value\nCafé,100\nNaïve,200"
    file_path.write_text(content, encoding='latin-1')

    from mcp_excel.formats.handlers import ParseOptions
    override = SheetOverride(header_rows=1)

    metas = loader.load_sheet(file_path, "override_encoding.csv", "Sheet1", "csv", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) == 2


def test_european_number_excel(temp_dir, loader):
    file_path = temp_dir / "european_excel.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(['Product', 'Price'])
    ws.append(['Widget', '1.234,56'])
    ws.append(['Gadget', '2.500,00'])
    ws.append(['Tool', '999,99'])

    wb.save(file_path)

    from mcp_excel.types import LocaleConfig
    override = SheetOverride(
        header_rows=1,
        locale=LocaleConfig(
            locale='de_DE',
            decimal_separator=',',
            thousands_separator='.',
            auto_detect=False
        )
    )
    metas = loader.load_sheet(file_path, "european_excel.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) == 3


def test_locale_config_manual_override(temp_dir, loader):
    file_path = temp_dir / "manual_locale.csv"
    content = """Product;Price
Widget;1.234,56
Gadget;2.500,00"""
    file_path.write_text(content, encoding='utf-8')

    from mcp_excel.types import LocaleConfig
    override = SheetOverride(
        header_rows=1,
        locale=LocaleConfig(
            locale='de_DE',
            decimal_separator=',',
            thousands_separator='.',
            currency_symbols=['€'],
            auto_detect=False
        )
    )
    metas = loader.load_sheet(file_path, "manual_locale.csv", "Sheet1", "csv", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) == 2


def test_mixed_locale_data(temp_dir, loader):
    file_path = temp_dir / "mixed_locale.csv"
    content = """Product,US_Price,EU_Price
Widget,1234.56,1.234;56
Gadget,2500.00,2.500;00"""
    file_path.write_text(content, encoding='utf-8')

    override = SheetOverride(auto_detect=True, header_rows=1)
    metas = loader.load_sheet(file_path, "mixed_locale.csv", "Sheet1", "csv", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) == 2


def test_fractional_excel_dates(temp_dir, loader):
    file_path = temp_dir / "fractional_dates.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(['Event', 'Timestamp'])
    ws.append(['Morning', 44927.5])
    ws.append(['Noon', 44927.5])
    ws.append(['Evening', 44927.75])

    ws['B2'].number_format = 'YYYY-MM-DD HH:MM'
    ws['B3'].number_format = 'YYYY-MM-DD HH:MM'
    ws['B4'].number_format = 'YYYY-MM-DD HH:MM'

    wb.save(file_path)

    override = SheetOverride(auto_detect=True, header_rows=1)
    metas = loader.load_sheet(file_path, "fractional_dates.xlsx", "Sheet", "excel", override)
    meta = metas[0]

    result = loader.conn.execute(f'SELECT * FROM "{meta.table_name}"').fetchall()
    assert len(result) == 3

import openpyxl
from pathlib import Path

def create_multi_table_fixture():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MultiTable"

    ws['A1'] = "Q1 Sales Report"

    ws['A3'] = "Product"
    ws['B3'] = "Revenue"
    ws['C3'] = "Units"

    ws['A4'] = "Widget A"
    ws['B4'] = 1500
    ws['C4'] = 50

    ws['A5'] = "Widget B"
    ws['B5'] = 2300
    ws['C5'] = 75

    ws['A6'] = "Widget C"
    ws['B6'] = 1800
    ws['C6'] = 60

    ws['A12'] = "Category"
    ws['B12'] = "Amount"
    ws['C12'] = "Department"

    ws['A13'] = "Marketing"
    ws['B13'] = 5000
    ws['C13'] = "Sales"

    ws['A14'] = "Operations"
    ws['B14'] = 8000
    ws['C14'] = "Ops"

    ws['A15'] = "Research"
    ws['B15'] = 12000
    ws['C15'] = "R&D"

    ws['A16'] = "Travel"
    ws['B16'] = 3500
    ws['C16'] = "All"

    ws_three = wb.create_sheet("ThreeTables")

    ws_three['A1'] = "North Region"
    ws_three['A2'] = "Sales"
    ws_three['B2'] = "Target"
    ws_three['A3'] = 100
    ws_three['B3'] = 120

    ws_three['A7'] = "South Region"
    ws_three['A8'] = "Sales"
    ws_three['B8'] = "Target"
    ws_three['A9'] = 150
    ws_three['B9'] = 140

    ws_three['A13'] = "East Region"
    ws_three['A14'] = "Sales"
    ws_three['B14'] = "Target"
    ws_three['A15'] = 200
    ws_three['B15'] = 180

    ws_single = wb.create_sheet("SingleTable")
    ws_single['A1'] = "Name"
    ws_single['B1'] = "Value"
    ws_single['A2'] = "Alpha"
    ws_single['B2'] = 100
    ws_single['A3'] = "Beta"
    ws_single['B3'] = 200

    fixtures_dir = Path(__file__).parent
    fixtures_dir.mkdir(exist_ok=True)
    output_path = fixtures_dir / "multi_table_test.xlsx"
    wb.save(output_path)
    print(f"Created fixture: {output_path}")

if __name__ == "__main__":
    create_multi_table_fixture()

import tempfile
from pathlib import Path
import pytest
import openpyxl
import time

from mcp_excel.structure_analyzer import ExcelAnalyzer, LRUCache


def test_lru_cache_basic():
    cache = LRUCache(maxsize=3)

    cache.put("key1", "value1")
    cache.put("key2", "value2")
    cache.put("key3", "value3")

    assert cache.get("key1") == "value1"
    assert cache.get("key2") == "value2"
    assert cache.get("key3") == "value3"
    assert len(cache) == 3


def test_lru_cache_eviction():
    cache = LRUCache(maxsize=2)

    cache.put("key1", "value1")
    cache.put("key2", "value2")
    cache.put("key3", "value3")

    assert cache.get("key1") is None
    assert cache.get("key2") == "value2"
    assert cache.get("key3") == "value3"
    assert len(cache) == 2


def test_lru_cache_update_access_order():
    cache = LRUCache(maxsize=2)

    cache.put("key1", "value1")
    cache.put("key2", "value2")

    cache.get("key1")

    cache.put("key3", "value3")

    assert cache.get("key1") == "value1"
    assert cache.get("key2") is None
    assert cache.get("key3") == "value3"


def test_lru_cache_overwrite():
    cache = LRUCache(maxsize=2)

    cache.put("key1", "value1")
    cache.put("key1", "new_value1")

    assert cache.get("key1") == "new_value1"
    assert len(cache) == 1


def test_lru_cache_clear():
    cache = LRUCache(maxsize=3)

    cache.put("key1", "value1")
    cache.put("key2", "value2")

    cache.clear()

    assert len(cache) == 0
    assert cache.get("key1") is None


def test_structure_analyzer_cache_size():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["A", "B"])
    ws.append(["1", "2"])

    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as f:
        wb.save(f.name)
        file_path = Path(f.name)

    analyzer = ExcelAnalyzer(cache_size=2)

    analyzer.analyze_structure(file_path, "Sheet")
    assert len(analyzer._cache) == 1

    analyzer.analyze_structure(file_path, "Sheet")
    assert len(analyzer._cache) == 1


def test_structure_analyzer_cache_hit():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["A", "B"])
    ws.append(["1", "2"])

    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as f:
        wb.save(f.name)
        file_path = Path(f.name)

    analyzer = ExcelAnalyzer()

    start_time = time.time()
    result1 = analyzer.analyze_structure(file_path, "Sheet")
    first_duration = time.time() - start_time

    start_time = time.time()
    result2 = analyzer.analyze_structure(file_path, "Sheet")
    second_duration = time.time() - start_time

    assert result1.data_start_row == result2.data_start_row
    assert second_duration < first_duration


def test_structure_analyzer_cache_mtime_invalidation():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["A", "B"])
    ws.append(["1", "2"])

    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as f:
        wb.save(f.name)
        file_path = Path(f.name)

    analyzer = ExcelAnalyzer()

    result1 = analyzer.analyze_structure(file_path, "Sheet")
    assert len(analyzer._cache) == 1

    time.sleep(0.1)

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    ws.append(["3", "4"])
    wb.save(file_path)

    result2 = analyzer.analyze_structure(file_path, "Sheet")

    assert len(analyzer._cache) == 2


def test_structure_analyzer_multiple_sheets():
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["A", "B"])

    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["C", "D"])

    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as f:
        wb.save(f.name)
        file_path = Path(f.name)

    analyzer = ExcelAnalyzer()

    analyzer.analyze_structure(file_path, "Sheet1")
    analyzer.analyze_structure(file_path, "Sheet2")

    assert len(analyzer._cache) == 2


def test_lru_cache_large_dataset():
    cache = LRUCache(maxsize=100)

    for i in range(200):
        cache.put(f"key{i}", f"value{i}")

    assert len(cache) == 100

    assert cache.get("key0") is None
    assert cache.get("key99") is None

    assert cache.get("key100") is not None
    assert cache.get("key199") is not None


def test_structure_analyzer_default_cache_size():
    analyzer = ExcelAnalyzer()

    assert analyzer._cache.maxsize == 128


def test_structure_analyzer_custom_cache_size():
    analyzer = ExcelAnalyzer(cache_size=256)

    assert analyzer._cache.maxsize == 256


def test_lru_cache_contains():
    cache = LRUCache(maxsize=2)

    cache.put("key1", "value1")

    assert "key1" in cache
    assert "key2" not in cache
